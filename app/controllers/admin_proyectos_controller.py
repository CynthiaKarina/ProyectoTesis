from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_required, current_user
from app.models.proyecto import Proyecto
from app.models.user import User
from app import db
from datetime import datetime, date
from app.utils.permissions import permission_required, any_permission_required, has_permission
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import or_
import os
from werkzeug.utils import secure_filename
# Importación de xhtml2pdf se realiza dentro de la función para evitar errores si no está instalado


admin_proyectos_bp = Blueprint('admin_proyectos', __name__)

@admin_proyectos_bp.route('/admin/proyectos')
@login_required
@permission_required('gestionar_proyectos')
def admin_proyectos():
    """Página principal de administración de proyectos"""
    try:
        # Obtener todos los proyectos
        proyectos = Proyecto.query.all()
        
        # Estadísticas dinámicas
        total_proyectos = len(proyectos)
        
        # Calcular estadísticas por estado
        estadisticas_estados = {}
        estados_disponibles = ['En Desarrollo', 'Completado', 'Suspendido', 'Cancelado', 'En Revision', 'Aprobado']
        
        for estado in estados_disponibles:
            estadisticas_estados[estado.lower().replace(' ', '_')] = len([p for p in proyectos if p.estatus == estado])
        
        # Estadísticas por tipo
        tipos_proyecto = set([p.tipo_proyecto for p in proyectos if p.tipo_proyecto])
        estadisticas_tipos = {tipo: len([p for p in proyectos if p.tipo_proyecto == tipo]) for tipo in tipos_proyecto}
        
        # Configuración de tarjetas de estadísticas
        stats_config = [
            {
                'key': 'total',
                'icon': 'fas fa-project-diagram',
                'label': 'Total Proyectos',
                'value': total_proyectos,
                'color': 'total'
            },
            {
                'key': 'en_desarrollo',
                'icon': 'fas fa-cogs',
                'label': 'En Desarrollo',
                'value': estadisticas_estados.get('en_desarrollo', 0),
                'color': 'info'
            },
            {
                'key': 'completado',
                'icon': 'fas fa-check-circle',
                'label': 'Completados',
                'value': estadisticas_estados.get('completado', 0),
                'color': 'success'
            },
            {
                'key': 'suspendido',
                'icon': 'fas fa-pause-circle',
                'label': 'Suspendidos',
                'value': estadisticas_estados.get('suspendido', 0),
                'color': 'warning'
            },
            {
                'key': 'cancelado',
                'icon': 'fas fa-times-circle',
                'label': 'Cancelados',
                'value': estadisticas_estados.get('cancelado', 0),
                'color': 'danger'
            },
            {
                'key': 'aprobado',
                'icon': 'fas fa-thumbs-up',
                'label': 'Aprobados',
                'value': estadisticas_estados.get('aprobado', 0),
                'color': 'success'
            }
        ]
        
        # Tipos de proyecto disponibles
        tipos_proyecto_disponibles = [
            'Investigación', 'Desarrollo', 'Innovación', 'Educativo',
            'Empresarial', 'Social', 'Tecnológico', 'Científico'
        ]
        
        return render_template('admin_proyectos.html', 
                             proyectos=proyectos,
                             total_proyectos=total_proyectos,
                             estadisticas_estados=estadisticas_estados,
                             estadisticas_tipos=estadisticas_tipos,
                             stats_config=stats_config,
                             tipos_proyecto_disponibles=tipos_proyecto_disponibles)
    except Exception as e:
        flash(f'Error al cargar la página de administración: {str(e)}', 'error')
        return redirect(url_for('home.index'))


@admin_proyectos_bp.route('/proyectos/nuevo', methods=['GET'])
@login_required
@permission_required('crear_proyecto')
def crear_proyecto_form():
    """Formulario simple para creación de proyectos (usuarios con permiso crear_proyecto)."""
    try:
        # Valores por defecto para selects
        estados_disponibles = ['En Desarrollo', 'Completado', 'Suspendido', 'Cancelado', 'En Revision', 'Aprobado']
        tipos_proyecto_disponibles = [
            'Investigación', 'Desarrollo', 'Innovación', 'Educativo',
            'Empresarial', 'Social', 'Tecnológico', 'Científico'
        ]
        # Opciones de usuarios para integrantes (filtrar por institución si existe)
        try:
            query = db.session.query(User)
            if hasattr(current_user, 'id_institucion') and current_user.id_institucion:
                query = query.filter(User.id_institucion == current_user.id_institucion)
            usuarios = query.order_by(User.nombre.asc()).limit(200).all()
            usuarios_opciones = [
                {
                    'id': u.id_usuario,
                    'nombre': f"{(u.nombre or '').strip()} {(u.apellido_paterno or '').strip()} {(u.apellido_materno or '').strip()}".strip() or u.username
                } for u in usuarios if hasattr(u, 'id_usuario')
            ]
        except Exception:
            usuarios_opciones = []

        return render_template(
            'proyecto_nuevo.html',
            estados_disponibles=estados_disponibles,
            tipos_proyecto=tipos_proyecto_disponibles,
            usuarios_opciones=usuarios_opciones
        )
    except Exception as e:
        flash(f'Error al cargar el formulario: {str(e)}', 'error')
        return redirect(url_for('home.index'))


@admin_proyectos_bp.route('/proyectos/nuevo', methods=['POST'])
@login_required
@permission_required('crear_proyecto')
def crear_proyecto_simple():
    """Procesa el formulario simple y crea un proyecto asociado al usuario actual."""
    try:
        nombre = (request.form.get('nombre_proyecto') or '').strip()
        if not nombre:
            flash('El nombre del proyecto es obligatorio.', 'error')
            return redirect(url_for('admin_proyectos.crear_proyecto_form'))
        if len(nombre) < 5 or len(nombre) > 300:
            flash('El nombre debe tener entre 5 y 300 caracteres.', 'warning')
            return redirect(url_for('admin_proyectos.crear_proyecto_form'))

        resumen = (request.form.get('resumen') or '').strip()
        tipo = (request.form.get('tipo_proyecto') or '').strip()
        estatus = (request.form.get('estatus') or 'En Desarrollo').strip()
        tipos_validos = ['Investigación', 'Desarrollo', 'Innovación', 'Educativo', 'Empresarial', 'Social', 'Tecnológico', 'Científico']
        estados_validos = ['En Desarrollo', 'Completado', 'Suspendido', 'Cancelado', 'En Revision', 'Aprobado']
        if tipo and tipo not in tipos_validos:
            flash('Tipo de proyecto inválido.', 'warning')
            return redirect(url_for('admin_proyectos.crear_proyecto_form'))
        if estatus not in estados_validos:
            flash('Estatus de proyecto inválido.', 'warning')
            return redirect(url_for('admin_proyectos.crear_proyecto_form'))
        adjunto = True if request.form.get('adjunto') in ['on', 'true', '1'] else False
        fecha_str = (request.form.get('fecha') or '').strip()
        fecha_val = None
        if fecha_str:
            try:
                fecha_val = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Formato de fecha inválido. Use YYYY-MM-DD', 'warning')

        # Verificar duplicado
        existente = Proyecto.query.filter_by(nombre_proyecto=nombre).first()
        if existente:
            flash('Ya existe un proyecto con ese nombre.', 'error')
            return redirect(url_for('admin_proyectos.crear_proyecto_form'))

        # Crear proyecto
        nuevo = Proyecto(
            nombre_proyecto=nombre,
            resumen=resumen,
            tipo_proyecto=tipo,
            estatus=estatus or 'En Desarrollo',
            adjunto=adjunto,
            fecha=fecha_val,
            fecha_creacion=datetime.utcnow(),
            fecha_modificacion=datetime.utcnow()
        )
        # Configurar visibilidad/aprobación
        try:
            from flask_login import current_user
            nuevo.owner_id = getattr(current_user, 'id_usuario', None)
            # Estudiante requiere aprobación; Investigador con rol director/docente/autor podría omitir
            nuevo.requiere_aprobacion = True
            from app.utils.permissions import get_user_role
            role = (get_user_role() or '').lower()
            if 'investigador' in role:
                # Si el creador es investigador, permitir marcar no aprobación inmediata si es responsable
                nuevo.requiere_aprobacion = False
                nuevo.publico = True
                nuevo.aprobado_por = getattr(current_user, 'id_usuario', None)
                nuevo.aprobado_en = datetime.utcnow()
        except Exception:
            pass
        db.session.add(nuevo)
        db.session.flush()

        # Asociar al usuario actual en tabla puente si existe
        try:
            from app.models.proyecto_usuario import ProyectoUsuario
            if hasattr(current_user, 'id_usuario') and current_user.id_usuario and nuevo.id_proyecto:
                vinculo = ProyectoUsuario(id_proyecto=nuevo.id_proyecto, id_usuario=current_user.id_usuario, rol_en_proyecto='Propietario')
                db.session.add(vinculo)
        except Exception:
            pass

        # Si requiere aprobación, crear solicitud de publicación
        try:
            if getattr(nuevo, 'requiere_aprobacion', True):
                from app.models.proyecto_request import ProyectoRequest
                pr = ProyectoRequest(id_proyecto=nuevo.id_proyecto, id_solicitante=current_user.id_usuario, tipo='publicacion', cambios='Creación inicial')
                db.session.add(pr)
        except Exception:
            pass

        # Integrantes adicionales seleccionados en el formulario
        try:
            from app.models.proyecto_usuario import ProyectoUsuario
            integrantes_ids = request.form.getlist('integrantes')
            if integrantes_ids:
                # Normalizar ids y evitar duplicar al propietario
                unique_ids = set()
                for raw_id in integrantes_ids:
                    try:
                        uid = int(raw_id)
                        if uid > 0 and uid != getattr(current_user, 'id_usuario', 0):
                            unique_ids.add(uid)
                    except ValueError:
                        continue
                for uid in unique_ids:
                    db.session.add(ProyectoUsuario(id_proyecto=nuevo.id_proyecto, id_usuario=uid, rol_en_proyecto='Integrante'))
        except Exception:
            pass

        # Audit: crear
        try:
            from app.models.proyecto_audit import ProyectoAudit
            db.session.add(ProyectoAudit(
                id_proyecto=nuevo.id_proyecto,
                id_usuario=getattr(current_user, 'id_usuario', None),
                accion='crear',
                detalles=f'Creado por usuario {getattr(current_user, "id_usuario", None)}'
            ))
        except Exception:
            pass

        db.session.commit()
        flash('Proyecto creado exitosamente.', 'success')
        return redirect(url_for('home.index'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear el proyecto: {str(e)}', 'error')
        return redirect(url_for('admin_proyectos.crear_proyecto_form'))


@admin_proyectos_bp.route('/mis-proyectos/pdf')
@login_required
def descargar_mis_proyectos_pdf():
    """Descargar en PDF los proyectos del usuario actual"""
    try:
        try:
            from xhtml2pdf import pisa
        except ImportError:
            flash('Generación de PDF no disponible. Instala xhtml2pdf (pip install xhtml2pdf) y reinicia.', 'warning')
            return redirect(url_for('home.index'))

        usuario_id = current_user.id_usuario if hasattr(current_user, 'id_usuario') else None
        if not usuario_id:
            flash('Usuario no autenticado.', 'error')
            return redirect(url_for('home.index'))

        # Obtener proyectos del usuario vía tabla puente si existe
        from app.models.proyecto_usuario import ProyectoUsuario
        proyectos = (db.session.query(Proyecto)
                     .join(ProyectoUsuario, ProyectoUsuario.id_proyecto == Proyecto.id_proyecto)
                     .filter(ProyectoUsuario.id_usuario == usuario_id)
                     .order_by(Proyecto.fecha_modificacion.desc())
                     .all())

        # Renderizar HTML simple
        html = render_template('proyectos_pdf.html', proyectos=proyectos, user=current_user)

        # Convertir a PDF
        result = io.BytesIO()
        pisa_status = pisa.CreatePDF(src=html, dest=result)
        if pisa_status.err:
            return jsonify({'success': False, 'message': 'Error al generar PDF'}), 500

        result.seek(0)
        response = make_response(result.read())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=mis_proyectos.pdf'
        return response
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al generar PDF: {str(e)}'}), 500


@admin_proyectos_bp.route('/proyectos/editar/<int:id_proyecto>', methods=['GET'])
@login_required
def editar_proyecto_form(id_proyecto):
    """Formulario de edición para propietarios del proyecto."""
    try:
        proyecto = Proyecto.query.get_or_404(id_proyecto)
        # Verificar que el usuario sea propietario
        from app.models.proyecto_usuario import ProyectoUsuario
        rel = ProyectoUsuario.query.filter_by(id_proyecto=id_proyecto, id_usuario=current_user.id_usuario).first()
        if not rel or (rel.rol_en_proyecto or '').lower() != 'propietario':
            flash('Solo el propietario puede editar este proyecto.', 'error')
            return redirect(url_for('home.index'))

        estados_disponibles = ['En Desarrollo', 'Completado', 'Suspendido', 'Cancelado', 'En Revision', 'Aprobado']
        tipos_proyecto_disponibles = [
            'Investigación', 'Desarrollo', 'Innovación', 'Educativo',
            'Empresarial', 'Social', 'Tecnológico', 'Científico'
        ]
        es_final = proyecto.estatus == 'Completado'
        # Historial de auditoría
        try:
            from app.models.proyecto_audit import ProyectoAudit
            audit_rows = (ProyectoAudit.query
                          .filter_by(id_proyecto=id_proyecto)
                          .order_by(ProyectoAudit.fecha.desc())
                          .limit(20)
                          .all())
        except Exception:
            audit_rows = []
        return render_template('proyecto_editar.html', proyecto=proyecto, estados_disponibles=estados_disponibles, tipos_proyecto=tipos_proyecto_disponibles, es_final=es_final, audit_rows=audit_rows)
    except Exception as e:
        flash(f'Error al cargar edición: {str(e)}', 'error')
        return redirect(url_for('home.index'))


@admin_proyectos_bp.route('/proyectos/editar/<int:id_proyecto>', methods=['POST'])
@login_required
def editar_proyecto_guardar(id_proyecto):
    """Guardar cambios (solo propietario y si no está finalizado)."""
    try:
        proyecto = Proyecto.query.get_or_404(id_proyecto)
        from app.models.proyecto_usuario import ProyectoUsuario
        rel = ProyectoUsuario.query.filter_by(id_proyecto=id_proyecto, id_usuario=current_user.id_usuario).first()
        if not rel or (rel.rol_en_proyecto or '').lower() != 'propietario':
            flash('Solo el propietario puede editar este proyecto.', 'error')
            return redirect(url_for('home.index'))
        if proyecto.estatus == 'Completado':
            flash('Este proyecto está finalizado y no se puede editar.', 'warning')
            return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))

        nombre = (request.form.get('nombre_proyecto') or '').strip()
        if not nombre or len(nombre) < 5 or len(nombre) > 300:
            flash('El nombre debe tener entre 5 y 300 caracteres.', 'warning')
            return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))

        tipos_validos = ['Investigación', 'Desarrollo', 'Innovación', 'Educativo', 'Empresarial', 'Social', 'Tecnológico', 'Científico']
        estados_validos = ['En Desarrollo', 'Completado', 'Suspendido', 'Cancelado', 'En Revision', 'Aprobado']

        tipo = (request.form.get('tipo_proyecto') or '').strip()
        estatus = (request.form.get('estatus') or 'En Desarrollo').strip()
        if tipo and tipo not in tipos_validos:
            flash('Tipo de proyecto inválido.', 'warning')
            return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))
        if estatus not in estados_validos:
            flash('Estatus de proyecto inválido.', 'warning')
            return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))

        proyecto.nombre_proyecto = nombre
        proyecto.resumen = (request.form.get('resumen') or '').strip()
        proyecto.tipo_proyecto = tipo
        proyecto.estatus = estatus
        proyecto.adjunto = True if request.form.get('adjunto') in ['on', 'true', '1'] else False

        fecha_str = (request.form.get('fecha') or '').strip()
        if fecha_str:
            try:
                proyecto.fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Formato de fecha inválido. Use YYYY-MM-DD', 'warning')

        proyecto.fecha_modificacion = datetime.utcnow()
        # Audit: editar (guardamos cambios relevantes)
        try:
            from app.models.proyecto_audit import ProyectoAudit
            db.session.add(ProyectoAudit(
                id_proyecto=proyecto.id_proyecto,
                id_usuario=getattr(current_user, 'id_usuario', None),
                accion='editar',
                detalles=f'Cambios en nombre/tipo/estatus/fecha por usuario {getattr(current_user, "id_usuario", None)}'
            ))
        except Exception:
            pass

        db.session.commit()
        flash('Cambios guardados correctamente.', 'success')
        return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al guardar cambios: {str(e)}', 'error')
        return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))


@admin_proyectos_bp.route('/proyectos/editar/<int:id_proyecto>/finalizar', methods=['POST'])
@login_required
def finalizar_proyecto(id_proyecto):
    """Marcar proyecto como última versión (bloquea edición)."""
    try:
        proyecto = Proyecto.query.get_or_404(id_proyecto)
        from app.models.proyecto_usuario import ProyectoUsuario
        rel = ProyectoUsuario.query.filter_by(id_proyecto=id_proyecto, id_usuario=current_user.id_usuario).first()
        if not rel or (rel.rol_en_proyecto or '').lower() != 'propietario':
            flash('Solo el propietario puede finalizar este proyecto.', 'error')
            return redirect(url_for('home.index'))
        proyecto.estatus = 'Completado'
        proyecto.fecha_modificacion = datetime.utcnow()
        # Audit: finalizar
        try:
            from app.models.proyecto_audit import ProyectoAudit
            db.session.add(ProyectoAudit(
                id_proyecto=proyecto.id_proyecto,
                id_usuario=getattr(current_user, 'id_usuario', None),
                accion='finalizar',
                detalles='Marcado como última versión'
            ))
        except Exception:
            pass

        db.session.commit()
        flash('Proyecto marcado como última versión. Ya no es editable.', 'success')
        return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al finalizar el proyecto: {str(e)}', 'error')
        return redirect(url_for('admin_proyectos.editar_proyecto_form', id_proyecto=id_proyecto))

@admin_proyectos_bp.route('/admin/proyectos/crear', methods=['POST'])
@login_required
@permission_required('crear_proyecto')
def crear_proyecto():
    """Crear un nuevo proyecto"""
    try:
        data = request.get_json()
        
        # Validar campos obligatorios
        if not data.get('nombre_proyecto'):
            return jsonify({
                'success': False,
                'message': 'El nombre del proyecto es obligatorio'
            }), 400
        
        # Verificar que no existe un proyecto con el mismo nombre
        proyecto_existente = Proyecto.query.filter_by(
            nombre_proyecto=data['nombre_proyecto']
        ).first()
        
        if proyecto_existente:
            return jsonify({
                'success': False,
                'message': 'Ya existe un proyecto con este nombre'
            }), 400
        
        # Convertir fecha si se proporciona
        fecha_proyecto = None
        if data.get('fecha'):
            try:
                fecha_proyecto = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
                }), 400
        
        # Crear nuevo proyecto
        nuevo_proyecto = Proyecto(
            nombre_proyecto=data['nombre_proyecto'],
            resumen=data.get('resumen', ''),
            adjunto=data.get('adjunto', False),
            fecha=fecha_proyecto,
            tipo_proyecto=data.get('tipo_proyecto', ''),
            estatus=data.get('estatus', 'En Desarrollo'),
            fecha_creacion=datetime.utcnow(),
            fecha_modificacion=datetime.utcnow()
        )
        
        db.session.add(nuevo_proyecto)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Proyecto creado exitosamente',
            'proyecto': nuevo_proyecto.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al crear el proyecto: {str(e)}'
        }), 500

@admin_proyectos_bp.route('/admin/proyectos/<int:id_proyecto>')
@login_required
@permission_required('gestionar_proyectos')
def obtener_proyecto(id_proyecto):
    """Obtener datos de un proyecto específico"""
    try:
        proyecto = Proyecto.query.get_or_404(id_proyecto)
        return jsonify(proyecto.to_dict())
    except Exception as e:
        return jsonify({
            'error': f'Error al obtener el proyecto: {str(e)}'
        }), 500

@admin_proyectos_bp.route('/admin/proyectos/editar/<int:id_proyecto>', methods=['PUT'])
@login_required
@permission_required('editar_proyecto')
def editar_proyecto(id_proyecto):
    """Editar un proyecto existente"""
    try:
        proyecto = Proyecto.query.get_or_404(id_proyecto)
        data = request.get_json()
        
        # Validar campos obligatorios
        if not data.get('nombre_proyecto'):
            return jsonify({
                'success': False,
                'message': 'El nombre del proyecto es obligatorio'
            }), 400
        
        # Verificar que no existe otro proyecto con el mismo nombre
        proyecto_existente = Proyecto.query.filter(
            Proyecto.nombre_proyecto == data['nombre_proyecto'],
            Proyecto.id_proyecto != id_proyecto
        ).first()
        
        if proyecto_existente:
            return jsonify({
                'success': False,
                'message': 'Ya existe otro proyecto con este nombre'
            }), 400
        
        # Convertir fecha si se proporciona
        if data.get('fecha'):
            try:
                proyecto.fecha = datetime.strptime(data['fecha'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
                }), 400
        
        # Actualizar campos
        proyecto.nombre_proyecto = data['nombre_proyecto']
        proyecto.resumen = data.get('resumen', proyecto.resumen)
        proyecto.adjunto = data.get('adjunto', proyecto.adjunto) 
        proyecto.tipo_proyecto = data.get('tipo_proyecto', proyecto.tipo_proyecto)
        proyecto.estatus = data.get('estatus', proyecto.estatus)
        proyecto.fecha_modificacion = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Proyecto actualizado exitosamente',
            'proyecto': proyecto.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al actualizar el proyecto: {str(e)}'
        }), 500

@admin_proyectos_bp.route('/admin/proyectos/eliminar/<int:id_proyecto>', methods=['DELETE'])
@login_required
@permission_required('eliminar_proyecto')
def eliminar_proyecto(id_proyecto):
    """Eliminar un proyecto"""
    try:
        proyecto = Proyecto.query.get_or_404(id_proyecto)
        
        nombre_proyecto = proyecto.nombre_proyecto
        db.session.delete(proyecto)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Proyecto "{nombre_proyecto}" eliminado exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al eliminar el proyecto: {str(e)}'
        }), 500

@admin_proyectos_bp.route('/admin/proyectos/cambiar-estado/<int:id_proyecto>', methods=['POST'])
@login_required
@permission_required('editar_proyecto')
def cambiar_estado_proyecto(id_proyecto):
    """Cambiar el estado de un proyecto"""
    try:
        proyecto = Proyecto.query.get_or_404(id_proyecto)
        data = request.get_json()
        
        nuevo_estado = data.get('estado')
        estados_validos = ['En Desarrollo', 'Completado', 'Suspendido', 'Cancelado', 'En Revision', 'Aprobado']
        
        if nuevo_estado not in estados_validos:
            return jsonify({
                'success': False,
                'message': 'Estado no válido'
            }), 400
        
        proyecto.estatus = nuevo_estado
        proyecto.fecha_modificacion = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado del proyecto cambiado a {nuevo_estado}',
            'nuevo_estado': nuevo_estado
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al cambiar el estado: {str(e)}'
        }), 500

@admin_proyectos_bp.route('/admin/proyectos/exportar')
@login_required
@permission_required('gestionar_proyectos')
def exportar_proyectos():
    """Exportar proyectos a Excel"""
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyectos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Encabezados
        headers = [
            'ID', 'Nombre del Proyecto', 'Resumen', 'Tipo de Proyecto',
            'Estado', 'Tiene Adjunto', 'Fecha del Proyecto', 'Fecha Creación', 'Fecha Modificación'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        proyectos = Proyecto.query.all()
        
        for row, proyecto in enumerate(proyectos, 2):
            data = [
                proyecto.id_proyecto,
                proyecto.nombre_proyecto,
                proyecto.resumen or '',
                proyecto.tipo_proyecto or '',
                proyecto.estatus,
                'Sí' if proyecto.adjunto else 'No',
                proyecto.fecha.strftime('%Y-%m-%d') if proyecto.fecha else '',
                proyecto.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if proyecto.fecha_creacion else '',
                proyecto.fecha_modificacion.strftime('%Y-%m-%d %H:%M:%S') if proyecto.fecha_modificacion else ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=proyectos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error al exportar proyectos: {str(e)}', 'error')
        return redirect(url_for('admin_proyectos.admin_proyectos'))

@admin_proyectos_bp.route('/admin/proyectos/plantilla')
@login_required
@permission_required('crear_proyecto')
def descargar_plantilla():
    """Descargar plantilla Excel para importar proyectos"""
    try:
        wb = Workbook()
        
        # Hoja 1: Plantilla de datos
        ws_data = wb.active
        ws_data.title = "Plantilla_Proyectos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        optional_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Encabezados con colores (rojo=obligatorio, azul=opcional)
        headers = [
            ('nombre_proyecto', 'Nombre del Proyecto', True),
            ('resumen', 'Resumen', False),
            ('tipo_proyecto', 'Tipo de Proyecto', False),
            ('estatus', 'Estado', False),
            ('adjunto', 'Tiene Adjunto', False),
            ('fecha', 'Fecha del Proyecto', False)
        ]
        
        for col, (field, header, required) in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = required_fill if required else optional_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fila explicativa
        explanations = [
            'Nombre único del proyecto',
            'Descripción detallada del proyecto',
            'Investigación/Desarrollo/Innovación/etc.',
            'En Desarrollo/Completado/Suspendido/etc.',
            'Sí/No - Si tiene archivos adjuntos',
            'Fecha en formato YYYY-MM-DD'
        ]
        
        for col, explanation in enumerate(explanations, 1):
            cell = ws_data.cell(row=2, column=col, value=explanation)
            cell.font = Font(italic=True, size=9)
            cell.border = border
        
        # Ejemplos de datos
        ejemplos = [
            ['Sistema de Gestión Académica', 'Desarrollo de un sistema web para gestión de estudiantes y cursos', 'Tecnológico', 'En Desarrollo', 'Sí', '2024-06-15'],
            ['Investigación en IA', 'Estudio sobre aplicaciones de inteligencia artificial en educación', 'Investigación', 'En Revision', 'No', '2024-08-20'],
            ['Programa de Innovación Social', 'Iniciativa para mejorar la calidad de vida comunitaria', 'Social', 'Aprobado', 'Sí', '2024-09-01']
        ]
        
        for row, ejemplo in enumerate(ejemplos, 3):
            for col, value in enumerate(ejemplo, 1):
                cell = ws_data.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Hoja 2: Instrucciones
        ws_instructions = wb.create_sheet("Instrucciones")
        instructions = [
            "📋 INSTRUCCIONES PARA IMPORTAR PROYECTOS",
            "",
            "🔴 CAMPOS OBLIGATORIOS:",
            "• nombre_proyecto: Nombre único del proyecto",
            "",
            "🔵 CAMPOS OPCIONALES:",
            "• resumen: Descripción detallada del proyecto",
            "• tipo_proyecto: Investigación, Desarrollo, Innovación, Educativo, Empresarial, Social, Tecnológico, Científico",
            "• estatus: En Desarrollo, Completado, Suspendido, Cancelado, En Revision, Aprobado",
            "• adjunto: Sí/No - Indica si el proyecto tiene archivos adjuntos",
            "• fecha: Fecha del proyecto en formato YYYY-MM-DD",
            "",
            "⚠️ VALIDACIONES:",
            "• Los nombres de proyecto deben ser únicos",
            "• Las fechas deben tener formato YYYY-MM-DD (ej: 2024-12-25)",
            "• Los estados deben ser uno de los valores permitidos",
            "• Los tipos deben ser uno de los valores sugeridos",
            "",
            "🚨 ERRORES COMUNES:",
            "• Nombres de proyecto duplicados",
            "• Formato incorrecto en fechas",
            "• Estados o tipos no válidos",
            "",
            "💡 CONSEJOS:",
            "• Use la plantilla de ejemplo como guía",
            "• Mantenga el formato de las columnas",
            "• Elimine las filas de ejemplo antes de importar",
            "• El campo 'adjunto' acepta: Sí, Si, Yes, True, 1 para verdadero"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row, column=1, value=instruction)
            if instruction.startswith(("📋", "🔴", "🔵", "⚠️", "🚨", "💡")):
                cell.font = Font(bold=True, size=12)
            elif instruction.startswith("•"):
                cell.font = Font(size=10)
        
        # Ajustar ancho de columnas
        for ws in [ws_data, ws_instructions]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=plantilla_proyectos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error al generar la plantilla: {str(e)}', 'error')
        return redirect(url_for('admin_proyectos.admin_proyectos'))

@admin_proyectos_bp.route('/admin/proyectos/importar', methods=['POST'])
@login_required
@permission_required('crear_proyecto')
def importar_proyectos():
    """Importar proyectos desde archivo Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'El archivo debe ser un Excel (.xlsx o .xls)'}), 400
        
        # Leer archivo Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Obtener encabezados
        headers = [cell.value for cell in ws[1]]
        
        # Validar encabezados obligatorios
        campos_obligatorios = ['nombre_proyecto']
        headers_lower = [h.lower() if h else '' for h in headers]
        
        for campo in campos_obligatorios:
            if campo not in headers_lower:
                return jsonify({
                    'success': False,
                    'message': f'Falta el campo obligatorio: {campo}'
                }), 400
        
        # Procesar filas
        proyectos_creados = []
        errores = []
        advertencias = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Saltar filas vacías
                continue
            
            try:
                # Crear diccionario con los datos de la fila
                data = {}
                for i, header in enumerate(headers):
                    if i < len(row) and header:
                        data[header.lower()] = row[i]
                
                # Validar campos obligatorios
                if not data.get('nombre_proyecto'):
                    errores.append(f'Fila {row_num}: Nombre del proyecto es obligatorio')
                    continue
                
                # Verificar que no existe un proyecto con el mismo nombre
                proyecto_existente = Proyecto.query.filter_by(
                    nombre_proyecto=str(data['nombre_proyecto']).strip()
                ).first()
                
                if proyecto_existente:
                    errores.append(f'Fila {row_num}: Ya existe un proyecto con el nombre "{data["nombre_proyecto"]}"')
                    continue
                
                # Convertir fecha si se proporciona
                fecha_proyecto = None
                if data.get('fecha'):
                    try:
                        if isinstance(data['fecha'], date):
                            fecha_proyecto = data['fecha']
                        else:
                            fecha_proyecto = datetime.strptime(str(data['fecha']), '%Y-%m-%d').date()
                    except ValueError:
                        advertencias.append(f'Fila {row_num}: Formato de fecha inválido, se omitirá')
                
                # Crear proyecto
                nuevo_proyecto = Proyecto(
                    nombre_proyecto=str(data['nombre_proyecto']).strip(),
                    resumen=str(data.get('resumen', '')).strip(),
                    tipo_proyecto=str(data.get('tipo_proyecto', '')).strip(),
                    estatus=str(data.get('estatus', 'En Desarrollo')).strip(),
                    adjunto=str(data.get('adjunto', 'No')).strip().lower() in ['sí', 'si', 'yes', 'true', '1'],
                    fecha=fecha_proyecto,
                    fecha_creacion=datetime.utcnow(),
                    fecha_modificacion=datetime.utcnow()
                )
                
                db.session.add(nuevo_proyecto)
                proyectos_creados.append(nuevo_proyecto.nombre_proyecto)
                
            except Exception as e:
                errores.append(f'Fila {row_num}: Error inesperado - {str(e)}')
        
        # Confirmar transacción si no hay errores críticos
        if proyectos_creados:
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Importación completada. {len(proyectos_creados)} proyectos creados.',
            'proyectos_creados': proyectos_creados,
            'errores': errores,
            'advertencias': advertencias
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al importar archivo: {str(e)}'
        }), 500

@admin_proyectos_bp.route('/admin/proyectos/buscar')
@login_required
@permission_required('gestionar_proyectos')
def buscar_proyectos():
    """Buscar proyectos para autocompletado"""
    try:
        query = request.args.get('q', '').strip()
        
        if not query:
            return jsonify({'success': True, 'proyectos': []})
        
        proyectos = Proyecto.query.filter(
            or_(
                Proyecto.nombre_proyecto.ilike(f'%{query}%'),
                Proyecto.resumen.ilike(f'%{query}%'),
                Proyecto.tipo_proyecto.ilike(f'%{query}%')
            )
        ).limit(10).all()
        
        return jsonify({
            'success': True,
            'proyectos': [{
                'id': proyecto.id_proyecto,
                'nombre': proyecto.nombre_proyecto,
                'resumen': proyecto.resumen_corto,
                'tipo': proyecto.tipo_proyecto,
                'estado': proyecto.estatus
            } for proyecto in proyectos]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error en la búsqueda: {str(e)}'
        }), 500 