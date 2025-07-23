from flask import Blueprint, render_template, request, jsonify, current_app, session, make_response
from app.models.institucion import Institucion
from app.models.tipo_institucion import Tipo_Institucion
from app import db
from flask_login import current_user, login_required
from datetime import datetime
from app.utils.permissions import permission_required
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

admin_instituciones_bp = Blueprint('admin_instituciones', __name__)

def get_current_user_name():
    """
    Función optimizada para obtener el nombre del usuario que inició sesión
    Prioriza: nombre_completo → nombre → username → sesión → fallback
    """
    try:
        current_app.logger.info(f"DEBUG: === OBTENIENDO NOMBRE DE USUARIO (Instituciones) ===")
        
        # Caso 1: Usuario autenticado con Flask-Login
        if current_user and current_user.is_authenticated:
            current_app.logger.info(f"DEBUG: Usuario autenticado: {current_user}")
            
            # Opción 1: Nombre completo (nombre + apellidos)
            if hasattr(current_user, 'nombre_completo') and current_user.nombre_completo.strip():
                result = current_user.nombre_completo.strip()
                current_app.logger.info(f"DEBUG: ✅ Usando nombre_completo: '{result}'")
                return result
            
            # Opción 2: Solo el nombre de pila
            if hasattr(current_user, 'nombre') and current_user.nombre and current_user.nombre.strip():
                result = current_user.nombre.strip()
                current_app.logger.info(f"DEBUG: ✅ Usando nombre: '{result}'")
                return result
            
            # Opción 3: Username como respaldo
            if hasattr(current_user, 'username') and current_user.username and current_user.username.strip():
                result = current_user.username.strip()
                current_app.logger.info(f"DEBUG: ✅ Usando username: '{result}'")
                return result
                
            current_app.logger.warning("DEBUG: Usuario autenticado pero sin nombre válido")
        
        # Fallback a datos de sesión Flask
        if 'nombre' in session and session['nombre'] and session['nombre'].strip():
            result = session['nombre'].strip()
            current_app.logger.info(f"DEBUG: ✅ Usando session['nombre']: '{result}'")
            return result
        
        if 'username' in session and session['username'] and session['username'].strip():
            result = session['username'].strip()
            current_app.logger.info(f"DEBUG: ✅ Usando session['username']: '{result}'")
            return result
        
        # Último recurso: Valor por defecto
        current_app.logger.warning("DEBUG: ⚠️ Usando valor por defecto: 'Sistema'")
        return 'Sistema'
        
    except Exception as e:
        current_app.logger.error(f"ERROR CRÍTICO en get_current_user_name: {str(e)}")
        return 'Sistema'

@admin_instituciones_bp.route('/admin/instituciones')
@login_required
@permission_required('gestionar_instituciones')
def index():
    """Página principal de administración de instituciones"""
    instituciones = Institucion.query.all()
    tipos_institucion = Tipo_Institucion.query.all()
    
    # Estadísticas para el dashboard
    estadisticas = {
        'total_instituciones': len(instituciones),
        'instituciones_activas': len([i for i in instituciones if i.activo]),
        'instituciones_inactivas': len([i for i in instituciones if not i.activo]),
        'tipos_disponibles': len(tipos_institucion),
    }
    
    return render_template('admin_instituciones.html', 
                         instituciones=instituciones, 
                         tipos_institucion=tipos_institucion,
                         estadisticas=estadisticas)

@admin_instituciones_bp.route('/admin/instituciones/<int:id>')
@login_required
@permission_required('gestionar_instituciones')
def get_institucion(id):
    """Obtener una institución específica por ID"""
    institucion = Institucion.query.get_or_404(id)
    return jsonify({
        'id_institucion': institucion.id_institucion,
        'nombre_institucion': institucion.nombre_institucion,
        'calle': institucion.calle,
        'colonia': institucion.colonia,
        'municipio': institucion.municipio,
        'estado': institucion.estado,
        'codigo_postal': institucion.codigo_postal,
        'telefono': institucion.telefono,
        'email': institucion.email,
        'activo': institucion.activo,
        'id_tipo_institucion': institucion.id_tipo_institucion,
        'fecha_registro': institucion.fecha_registro.strftime('%Y-%m-%d %H:%M:%S') if hasattr(institucion.fecha_registro, 'strftime') and institucion.fecha_registro else str(institucion.fecha_registro) if institucion.fecha_registro else None,
        'tipo_institucion_nombre': institucion.tipo_institucion.nombre_tipo_institucion if institucion.tipo_institucion else None
    })

@admin_instituciones_bp.route('/admin/instituciones/agregar', methods=['POST'])
@login_required
@permission_required('crear_institucion')
def agregar_institucion():
    """Crear una nueva institución"""
    try:
        current_app.logger.info("=== INICIANDO CREACIÓN DE INSTITUCIÓN ===")
        
        # Obtener datos del formulario o JSON
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        
        nombre_institucion = data.get('nombre_institucion', '').strip()
        current_app.logger.info(f"📝 Nombre de la institución: '{nombre_institucion}'")
        
        # Validaciones obligatorias
        if not nombre_institucion:
            current_app.logger.error("❌ Nombre de institución vacío")
            return jsonify({'error': 'El nombre de la institución es obligatorio'}), 400
        
        # Verificar si ya existe una institución con ese nombre
        institucion_existente = Institucion.query.filter_by(nombre_institucion=nombre_institucion).first()
        if institucion_existente:
            current_app.logger.warning(f"⚠️ Institución ya existe: {nombre_institucion}")
            return jsonify({'error': f'Ya existe una institución con el nombre "{nombre_institucion}"'}), 400
        
        # Validar tipo de institución
        id_tipo_institucion = data.get('id_tipo_institucion')
        if id_tipo_institucion:
            tipo_institucion = Tipo_Institucion.query.get(id_tipo_institucion)
            if not tipo_institucion:
                return jsonify({'error': 'Tipo de institución no válido'}), 400
        else:
            # Buscar un tipo por defecto o el primero disponible
            tipo_default = Tipo_Institucion.query.first()
            if tipo_default:
                id_tipo_institucion = tipo_default.id_tipo_institucion
            else:
                return jsonify({'error': 'No hay tipos de institución disponibles'}), 400
        
        # Crear nueva institución
        nueva_institucion = Institucion(
            nombre_institucion=nombre_institucion,
            calle=data.get('calle', '').strip() or None,
            colonia=data.get('colonia', '').strip() or None,
            municipio=data.get('municipio', '').strip() or None,
            estado=data.get('estado', '').strip() or None,
            codigo_postal=data.get('codigo_postal', '').strip() or None,
            telefono=data.get('telefono', '').strip() or None,
            email=data.get('email', '').strip() or None,
            activo=data.get('activo', True),
            id_tipo_institucion=id_tipo_institucion,
            fecha_registro=datetime.utcnow()
        )
        
        current_app.logger.info(f"✅ Institución creada exitosamente:")
        current_app.logger.info(f"   - Nombre: '{nueva_institucion.nombre_institucion}'")
        current_app.logger.info(f"   - Tipo ID: {nueva_institucion.id_tipo_institucion}")
        
        # Guardar en la base de datos
        db.session.add(nueva_institucion)
        db.session.commit()
        
        current_app.logger.info("🎉 Institución guardada en base de datos exitosamente")
        return jsonify({
            'mensaje': f'Institución "{nueva_institucion.nombre_institucion}" creada exitosamente',
            'institucion': {
                'id_institucion': nueva_institucion.id_institucion,
                'nombre_institucion': nueva_institucion.nombre_institucion,
                'activo': nueva_institucion.activo,
                'fecha_registro': nueva_institucion.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')
            }
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"💥 ERROR CRÍTICO en agregar_institucion: {str(e)}")
        return jsonify({'error': f'Error al crear institución: {str(e)}'}), 500

@admin_instituciones_bp.route('/admin/instituciones/editar/<int:id>', methods=['PUT'])
@login_required
@permission_required('editar_institucion')
def editar_institucion(id):
    """Editar una institución existente"""
    try:
        institucion = Institucion.query.get_or_404(id)
        data = request.get_json()
        
        # Validar nombre requerido
        nuevo_nombre = data.get('nombre_institucion', '').strip()
        if not nuevo_nombre:
            return jsonify({'error': 'El nombre de la institución es obligatorio'}), 400
        
        # Verificar duplicados (excepto la misma institución)
        institucion_existente = Institucion.query.filter(
            Institucion.nombre_institucion == nuevo_nombre,
            Institucion.id_institucion != id
        ).first()
        if institucion_existente:
            return jsonify({'error': f'Ya existe otra institución con el nombre "{nuevo_nombre}"'}), 400
        
        # Actualizar campos
        institucion.nombre_institucion = nuevo_nombre
        institucion.calle = data.get('calle', '').strip() or None
        institucion.colonia = data.get('colonia', '').strip() or None
        institucion.municipio = data.get('municipio', '').strip() or None
        institucion.estado = data.get('estado', '').strip() or None
        institucion.codigo_postal = data.get('codigo_postal', '').strip() or None
        institucion.telefono = data.get('telefono', '').strip() or None
        institucion.email = data.get('email', '').strip() or None
        
        # Manejar estado activo
        if 'activo' in data:
            institucion.activo = data.get('activo', True)
        
        # Validar y actualizar tipo de institución
        if 'id_tipo_institucion' in data and data['id_tipo_institucion']:
            tipo_institucion = Tipo_Institucion.query.get(data['id_tipo_institucion'])
            if tipo_institucion:
                institucion.id_tipo_institucion = data['id_tipo_institucion']
            else:
                return jsonify({'error': 'Tipo de institución no válido'}), 400
        
        db.session.commit()
        return jsonify({
            'mensaje': f'Institución "{institucion.nombre_institucion}" actualizada exitosamente',
            'institucion': {
                'id_institucion': institucion.id_institucion,
                'nombre_institucion': institucion.nombre_institucion,
                'activo': institucion.activo
            }
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error en editar_institucion: {str(e)}")
        return jsonify({'error': f'Error al actualizar institución: {str(e)}'}), 500

@admin_instituciones_bp.route('/admin/instituciones/eliminar/<int:id>', methods=['DELETE'])
@login_required
@permission_required('eliminar_institucion')
def eliminar_institucion(id):
    """Eliminar una institución"""
    try:
        institucion = Institucion.query.get_or_404(id)
        
        # Verificar si la institución tiene usuarios asignados
        if hasattr(institucion, 'usuarios') and institucion.usuarios:
            usuarios_count = len(institucion.usuarios)
            return jsonify({
                'error': f'No se puede eliminar la institución porque tiene {usuarios_count} usuario(s) asignado(s)'
            }), 400
        
        # Verificar si la institución tiene laboratorios asignados
        if hasattr(institucion, 'laboratorios') and institucion.laboratorios:
            laboratorios_count = len(institucion.laboratorios)
            return jsonify({
                'error': f'No se puede eliminar la institución porque tiene {laboratorios_count} laboratorio(s) asignado(s)'
            }), 400
        
        institucion_nombre = institucion.nombre_institucion  # Guardar para el mensaje
        db.session.delete(institucion)
        db.session.commit()
        return jsonify({'mensaje': f'Institución "{institucion_nombre}" eliminada exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error en eliminar_institucion: {str(e)}")
        return jsonify({'error': f'Error al eliminar institución: {str(e)}'}), 500

@admin_instituciones_bp.route('/admin/instituciones/toggle-status/<int:id>', methods=['POST'])
@login_required
@permission_required('cambiar_estado_institucion')
def toggle_status(id):
    """Cambiar el estado activo/inactivo de una institución"""
    try:
        institucion = Institucion.query.get_or_404(id)
        institucion.activo = not institucion.activo
        db.session.commit()
        
        estado_texto = "activada" if institucion.activo else "desactivada"
        return jsonify({
            'mensaje': f'Institución "{institucion.nombre_institucion}" {estado_texto} exitosamente',
            'activo': institucion.activo
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al cambiar estado: {str(e)}'}), 500

@admin_instituciones_bp.route('/admin/instituciones/exportar')
@login_required
@permission_required('exportar_instituciones')
def exportar_instituciones():
    """Exportar instituciones a Excel con formato mejorado"""
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Instituciones"
        
        # Estilos para el encabezado
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezados
        headers = [
            'ID', 'Nombre de la Institución', 'Tipo de Institución', 
            'Dirección Completa', 'Teléfono', 'Email', 'Estado', 
            'Fecha de Registro'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Obtener datos de instituciones con tipos
        instituciones = db.session.query(Institucion, Tipo_Institucion).outerjoin(
            Tipo_Institucion, Institucion.id_tipo_institucion == Tipo_Institucion.id_tipo_institucion
        ).all()
        
        # Escribir datos
        for row, (institucion, tipo_institucion) in enumerate(instituciones, 2):
            # Construir dirección completa
            direccion_parts = []
            if institucion.calle:
                direccion_parts.append(institucion.calle)
            if institucion.colonia:
                direccion_parts.append(f"Col. {institucion.colonia}")
            if institucion.municipio:
                direccion_parts.append(institucion.municipio)
            if institucion.estado:
                direccion_parts.append(institucion.estado)
            if institucion.codigo_postal:
                direccion_parts.append(f"C.P. {institucion.codigo_postal}")
            
            direccion_completa = ", ".join(direccion_parts) if direccion_parts else "Sin dirección"
            
            # Formatear fecha
            fecha_registro_str = 'Sin fecha'
            if institucion.fecha_registro:
                try:
                    fecha_registro_str = institucion.fecha_registro.strftime('%d/%m/%Y %H:%M')
                except:
                    fecha_registro_str = str(institucion.fecha_registro)
            
            data = [
                institucion.id_institucion,
                institucion.nombre_institucion,
                tipo_institucion.nombre_tipo_institucion if tipo_institucion else 'Sin tipo',
                direccion_completa,
                institucion.telefono or 'Sin teléfono',
                institucion.email or 'Sin email',
                'Activa' if institucion.activo else 'Inactiva',
                fecha_registro_str
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                # Estilo alternado para filas
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=instituciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Error en exportar_instituciones: {str(e)}")
        return jsonify({'error': f'Error al exportar instituciones: {str(e)}'}), 500

@admin_instituciones_bp.route('/admin/instituciones/exportar-csv')
@login_required
@permission_required('exportar_instituciones')
def exportar_instituciones_csv():
    """Exportar instituciones a CSV"""
    try:
        # Crear CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow([
            'ID', 'Nombre de la Institución', 'Tipo de Institución',
            'Dirección Completa', 'Teléfono', 'Email', 'Estado',
            'Fecha de Registro'
        ])
        
        # Obtener instituciones con tipos
        instituciones = db.session.query(Institucion, Tipo_Institucion).outerjoin(
            Tipo_Institucion, Institucion.id_tipo_institucion == Tipo_Institucion.id_tipo_institucion
        ).all()
        
        # Escribir datos
        for institucion, tipo_institucion in instituciones:
            # Construir dirección completa
            direccion_parts = []
            if institucion.calle:
                direccion_parts.append(institucion.calle)
            if institucion.colonia:
                direccion_parts.append(f"Col. {institucion.colonia}")
            if institucion.municipio:
                direccion_parts.append(institucion.municipio)
            if institucion.estado:
                direccion_parts.append(institucion.estado)
            if institucion.codigo_postal:
                direccion_parts.append(f"C.P. {institucion.codigo_postal}")
            
            direccion_completa = ", ".join(direccion_parts) if direccion_parts else "Sin dirección"
            
            # Formatear fecha
            fecha_registro_str = 'Sin fecha'
            if institucion.fecha_registro:
                try:
                    fecha_registro_str = institucion.fecha_registro.strftime('%d/%m/%Y %H:%M')
                except:
                    fecha_registro_str = str(institucion.fecha_registro)
            
            writer.writerow([
                institucion.id_institucion,
                institucion.nombre_institucion,
                tipo_institucion.nombre_tipo_institucion if tipo_institucion else 'Sin tipo',
                direccion_completa,
                institucion.telefono or 'Sin teléfono',
                institucion.email or 'Sin email',
                'Activa' if institucion.activo else 'Inactiva',
                fecha_registro_str
            ])
        
        # Preparar respuesta
        output.seek(0)
        csv_content = output.getvalue()
        
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=instituciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"Error en exportar_instituciones_csv: {str(e)}")
        return jsonify({'error': f'Error al exportar instituciones: {str(e)}'}), 500 