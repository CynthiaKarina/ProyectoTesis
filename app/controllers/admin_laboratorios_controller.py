from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_required, current_user
from app.models.laboratorio import Laboratorio
from app.models.area import Area
from app.models.institucion import Institucion
from app.models.tipo_laboratorio import Tipo_Laboratorio
from app.models.user import User
from app.models.laboratorio_imagen import LaboratorioImagen
from app import db
from datetime import datetime
from app.utils.permissions import permission_required, any_permission_required
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import or_
import os
from werkzeug.utils import secure_filename

admin_laboratorios_bp = Blueprint('admin_laboratorios', __name__)

@admin_laboratorios_bp.route('/admin/laboratorios')
@login_required
@permission_required('gestionar_laboratorios')
def admin_laboratorios():
    """Página principal de administración de laboratorios"""
    try:
        # Obtener datos con relaciones
        laboratorios = db.session.query(Laboratorio).options(
            db.joinedload(Laboratorio.institucion),
            db.joinedload(Laboratorio.area),
            db.joinedload(Laboratorio.tipo_laboratorio),
            db.joinedload(Laboratorio.encargado)
        ).all()
        
        # Datos para filtros y formularios
        areas = Area.query.all()
        instituciones = Institucion.query.all()
        tipos_laboratorio = Tipo_Laboratorio.query.all()
        usuarios = User.query.filter(User.activo == True).all()
        
        # Estadísticas
        total_laboratorios = len(laboratorios)
        laboratorios_activos = len([lab for lab in laboratorios if lab.disponibilidad == 'Activo'])
        laboratorios_mantenimiento = len([lab for lab in laboratorios if lab.disponibilidad == 'Mantenimiento'])
        
        return render_template('admin_laboratorios.html', 
                             laboratorios=laboratorios,
                             areas=areas,
                             instituciones=instituciones,
                             tipos_laboratorio=tipos_laboratorio,
                             usuarios=usuarios,
                             total_laboratorios=total_laboratorios,
                             laboratorios_activos=laboratorios_activos,
                             laboratorios_mantenimiento=laboratorios_mantenimiento)
    except Exception as e:
        flash(f'Error al cargar la página de administración: {str(e)}', 'error')
        return redirect(url_for('home.index'))

@admin_laboratorios_bp.route('/admin/laboratorios/crear', methods=['POST'])
@login_required
@permission_required('crear_laboratorio')
def crear_laboratorio():
    """Crear un nuevo laboratorio"""
    try:
        data = request.get_json()
        
        # Validar campos obligatorios
        campos_obligatorios = ['nombre_laboratorio', 'id_institucion', 'id_tipo_laboratorio']
        for campo in campos_obligatorios:
            if not data.get(campo):
                return jsonify({
                    'success': False,
                    'message': f'El campo {campo.replace("_", " ").title()} es obligatorio'
                }), 400
        
        # Verificar que no existe un laboratorio con el mismo nombre en la misma institución
        laboratorio_existente = Laboratorio.query.filter_by(
            nombre_laboratorio=data['nombre_laboratorio'],
            id_institucion=data['id_institucion']
        ).first()
        
        if laboratorio_existente:
            return jsonify({
                'success': False,
                'message': 'Ya existe un laboratorio con este nombre en la institución seleccionada'
            }), 400
        
        # Crear nuevo laboratorio
        nuevo_laboratorio = Laboratorio(
            nombre_laboratorio=data['nombre_laboratorio'],
            descripcion=data.get('descripcion', ''),
            disponibilidad=data.get('disponibilidad', 'Activo'),
            capacidad=int(data.get('capacidad', 20)),
            horario=data.get('horario', ''),
            ubicacion=data.get('ubicacion', ''),
            telefono=data.get('telefono', ''),
            email_contacto=data.get('email_contacto', ''),
            superficie_m2=float(data.get('superficie_m2', 0)) if data.get('superficie_m2') else None,
            equipamiento=data.get('equipamiento', ''),
            normas_uso=data.get('normas_uso', ''),
            requiere_capacitacion=data.get('requiere_capacitacion', False),
            id_institucion=int(data['id_institucion']),
            id_area=int(data['id_area']) if data.get('id_area') else None,
            id_tipo_laboratorio=int(data['id_tipo_laboratorio']),
            id_encargado=int(data['id_encargado']) if data.get('id_encargado') else None,
            fecha_creacion=datetime.utcnow(),
            fecha_modificacion=datetime.utcnow()
        )
        
        db.session.add(nuevo_laboratorio)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Laboratorio creado exitosamente',
            'laboratorio': nuevo_laboratorio.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al crear el laboratorio: {str(e)}'
        }), 500

@admin_laboratorios_bp.route('/admin/laboratorios/editar/<int:id_laboratorio>', methods=['PUT'])
@login_required
@permission_required('editar_laboratorio')
def editar_laboratorio(id_laboratorio):
    """Editar un laboratorio existente"""
    try:
        laboratorio = Laboratorio.query.get_or_404(id_laboratorio)
        data = request.get_json()
        
        # Validar campos obligatorios
        campos_obligatorios = ['nombre_laboratorio', 'id_institucion', 'id_tipo_laboratorio']
        for campo in campos_obligatorios:
            if not data.get(campo):
                return jsonify({
                    'success': False,
                    'message': f'El campo {campo.replace("_", " ").title()} es obligatorio'
                }), 400
        
        # Verificar que no existe otro laboratorio con el mismo nombre en la misma institución
        laboratorio_existente = Laboratorio.query.filter(
            Laboratorio.nombre_laboratorio == data['nombre_laboratorio'],
            Laboratorio.id_institucion == data['id_institucion'],
            Laboratorio.id_laboratorio != id_laboratorio
        ).first()
        
        if laboratorio_existente:
            return jsonify({
                'success': False,
                'message': 'Ya existe otro laboratorio con este nombre en la institución seleccionada'
            }), 400
        
        # Actualizar campos
        laboratorio.nombre_laboratorio = data['nombre_laboratorio']
        laboratorio.descripcion = data.get('descripcion', laboratorio.descripcion)
        laboratorio.disponibilidad = data.get('disponibilidad', laboratorio.disponibilidad)
        laboratorio.capacidad = int(data.get('capacidad', laboratorio.capacidad))
        laboratorio.horario = data.get('horario', laboratorio.horario)
        laboratorio.ubicacion = data.get('ubicacion', laboratorio.ubicacion)
        laboratorio.telefono = data.get('telefono', laboratorio.telefono)
        laboratorio.email_contacto = data.get('email_contacto', laboratorio.email_contacto)
        laboratorio.superficie_m2 = float(data.get('superficie_m2', 0)) if data.get('superficie_m2') else laboratorio.superficie_m2
        laboratorio.equipamiento = data.get('equipamiento', laboratorio.equipamiento)
        laboratorio.normas_uso = data.get('normas_uso', laboratorio.normas_uso)
        laboratorio.requiere_capacitacion = data.get('requiere_capacitacion', laboratorio.requiere_capacitacion)
        laboratorio.id_institucion = int(data['id_institucion'])
        laboratorio.id_area = int(data['id_area']) if data.get('id_area') else None
        laboratorio.id_tipo_laboratorio = int(data['id_tipo_laboratorio'])
        laboratorio.id_encargado = int(data['id_encargado']) if data.get('id_encargado') else None
        laboratorio.fecha_modificacion = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Laboratorio actualizado exitosamente',
            'laboratorio': laboratorio.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al actualizar el laboratorio: {str(e)}'
        }), 500

@admin_laboratorios_bp.route('/admin/laboratorios/eliminar/<int:id_laboratorio>', methods=['DELETE'])
@login_required
@permission_required('eliminar_laboratorio')
def eliminar_laboratorio(id_laboratorio):
    """Eliminar un laboratorio"""
    try:
        laboratorio = Laboratorio.query.get_or_404(id_laboratorio)
        
        # Verificar si el laboratorio tiene solicitudes asociadas
        if hasattr(laboratorio, 'solicitudes') and laboratorio.solicitudes:
            return jsonify({
                'success': False,
                'message': 'No se puede eliminar el laboratorio porque tiene solicitudes asociadas'
            }), 400
        
        # Eliminar imágenes asociadas
        for imagen in laboratorio.imagenes:
            db.session.delete(imagen)
        
        nombre_laboratorio = laboratorio.nombre_laboratorio
        db.session.delete(laboratorio)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Laboratorio "{nombre_laboratorio}" eliminado exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al eliminar el laboratorio: {str(e)}'
        }), 500

@admin_laboratorios_bp.route('/admin/laboratorios/cambiar-estado/<int:id_laboratorio>', methods=['POST'])
@login_required
@permission_required('editar_laboratorio')
def cambiar_estado_laboratorio(id_laboratorio):
    """Cambiar el estado de disponibilidad de un laboratorio"""
    try:
        laboratorio = Laboratorio.query.get_or_404(id_laboratorio)
        data = request.get_json()
        
        nuevo_estado = data.get('estado')
        estados_validos = ['Activo', 'Inactivo', 'Mantenimiento', 'Reservado']
        
        if nuevo_estado not in estados_validos:
            return jsonify({
                'success': False,
                'message': 'Estado no válido'
            }), 400
        
        laboratorio.disponibilidad = nuevo_estado
        laboratorio.fecha_modificacion = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado del laboratorio cambiado a {nuevo_estado}',
            'nuevo_estado': nuevo_estado
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al cambiar el estado: {str(e)}'
        }), 500

@admin_laboratorios_bp.route('/admin/laboratorios/exportar')
@login_required
@permission_required('gestionar_laboratorios')
def exportar_laboratorios():
    """Exportar laboratorios a Excel"""
    try:
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Laboratorios"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Descripción', 'Disponibilidad', 'Capacidad',
            'Horario', 'Ubicación', 'Teléfono', 'Email', 'Superficie (m²)',
            'Equipamiento', 'Requiere Capacitación', 'Institución', 'Área',
            'Tipo Laboratorio', 'Encargado', 'Fecha Creación'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        laboratorios = db.session.query(Laboratorio).options(
            db.joinedload(Laboratorio.institucion),
            db.joinedload(Laboratorio.area),
            db.joinedload(Laboratorio.tipo_laboratorio),
            db.joinedload(Laboratorio.encargado)
        ).all()
        
        for row, lab in enumerate(laboratorios, 2):
            data = [
                lab.id_laboratorio,
                lab.nombre_laboratorio,
                lab.descripcion or '',
                lab.disponibilidad,
                lab.capacidad,
                lab.horario or '',
                lab.ubicacion or '',
                lab.telefono or '',
                lab.email_contacto or '',
                float(lab.superficie_m2) if lab.superficie_m2 else '',
                lab.equipamiento or '',
                'Sí' if lab.requiere_capacitacion else 'No',
                lab.institucion.nombre_institucion if lab.institucion else '',
                lab.area.nombre_area if lab.area else '',
                lab.tipo_laboratorio.nombre_tipo_laboratorio if lab.tipo_laboratorio else '',
                lab.encargado.nombre_completo if lab.encargado else '',
                lab.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if lab.fecha_creacion else ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=laboratorios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error al exportar laboratorios: {str(e)}', 'error')
        return redirect(url_for('admin_laboratorios.admin_laboratorios'))

@admin_laboratorios_bp.route('/admin/laboratorios/plantilla')
@login_required
@permission_required('crear_laboratorio')
def descargar_plantilla():
    """Descargar plantilla Excel para importar laboratorios"""
    try:
        wb = Workbook()
        
        # Hoja 1: Plantilla de datos
        ws_data = wb.active
        ws_data.title = "Plantilla_Laboratorios"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        optional_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Encabezados con colores (rojo=obligatorio, azul=opcional)
        headers = [
            ('nombre_laboratorio', 'Nombre del Laboratorio', True),
            ('descripcion', 'Descripción', False),
            ('disponibilidad', 'Disponibilidad', False),
            ('capacidad', 'Capacidad', False),
            ('horario', 'Horario', False),
            ('ubicacion', 'Ubicación', False),
            ('telefono', 'Teléfono', False),
            ('email_contacto', 'Email de Contacto', False),
            ('superficie_m2', 'Superficie (m²)', False),
            ('equipamiento', 'Equipamiento', False),
            ('normas_uso', 'Normas de Uso', False),
            ('requiere_capacitacion', 'Requiere Capacitación', False),
            ('id_institucion', 'ID Institución', True),
            ('id_area', 'ID Área', False),
            ('id_tipo_laboratorio', 'ID Tipo Laboratorio', True),
            ('id_encargado', 'ID Encargado', False)
        ]
        
        for col, (field, header, required) in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = required_fill if required else optional_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fila explicativa
        explanations = [
            'Nombre único del laboratorio',
            'Descripción detallada (opcional)',
            'Activo/Inactivo/Mantenimiento/Reservado',
            'Número de personas (por defecto: 20)',
            'Horario de funcionamiento',
            'Ubicación física',
            'Teléfono de contacto',
            'Email de contacto',
            'Superficie en metros cuadrados',
            'Lista de equipos separados por comas',
            'Normas y reglamentos',
            'Sí/No - Si requiere capacitación',
            'ID de la institución (obligatorio)',
            'ID del área académica',
            'ID del tipo de laboratorio (obligatorio)',
            'ID del usuario encargado'
        ]
        
        for col, explanation in enumerate(explanations, 1):
            cell = ws_data.cell(row=2, column=col, value=explanation)
            cell.font = Font(italic=True, size=9)
            cell.border = border
        
        # Ejemplos de datos
        ejemplos = [
            ['Laboratorio de Química Analítica', 'Laboratorio equipado para análisis químicos', 'Activo', 25, 
             'Lunes a Viernes 8:00-18:00', 'Edificio A, Piso 2', '555-0123', 'quimica@universidad.edu', 
             45.5, 'Espectrofotómetro, Balanza analítica, Campana extractora', 'Uso obligatorio de EPP', 'Sí', 1, 1, 1, 2],
            ['Laboratorio de Física', 'Laboratorio para experimentos de física', 'Activo', 30, 
             'Lunes a Viernes 9:00-17:00', 'Edificio B, Piso 1', '555-0124', 'fisica@universidad.edu', 
             60.0, 'Osciloscopio, Generador de funciones, Multímetros', 'Prohibido alimentos y bebidas', 'No', 1, 2, 2, 3],
            ['Laboratorio de Biología', 'Laboratorio para estudios biológicos', 'Mantenimiento', 20, 
             'Lunes a Viernes 8:00-16:00', 'Edificio C, Piso 3', '555-0125', 'biologia@universidad.edu', 
             40.0, 'Microscopios, Incubadora, Autoclave', 'Desinfección obligatoria', 'Sí', 2, 3, 3, 4]
        ]
        
        for row, ejemplo in enumerate(ejemplos, 3):
            for col, value in enumerate(ejemplo, 1):
                cell = ws_data.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Hoja 2: Instrucciones
        ws_instructions = wb.create_sheet("Instrucciones")
        instructions = [
            "📋 INSTRUCCIONES PARA IMPORTAR LABORATORIOS",
            "",
            "🔴 CAMPOS OBLIGATORIOS:",
            "• nombre_laboratorio: Nombre único del laboratorio",
            "• id_institucion: ID de la institución (ver hoja Referencias)",
            "• id_tipo_laboratorio: ID del tipo de laboratorio (ver hoja Referencias)",
            "",
            "🔵 CAMPOS OPCIONALES:",
            "• descripcion: Descripción detallada del laboratorio",
            "• disponibilidad: Estado (Activo/Inactivo/Mantenimiento/Reservado)",
            "• capacidad: Número máximo de personas (por defecto: 20)",
            "• horario: Horario de funcionamiento",
            "• ubicacion: Ubicación física del laboratorio",
            "• telefono: Teléfono de contacto",
            "• email_contacto: Email de contacto",
            "• superficie_m2: Superficie en metros cuadrados",
            "• equipamiento: Lista de equipos separados por comas",
            "• normas_uso: Normas y reglamentos del laboratorio",
            "• requiere_capacitacion: Sí/No - Si requiere capacitación previa",
            "• id_area: ID del área académica",
            "• id_encargado: ID del usuario encargado",
            "",
            "⚠️ VALIDACIONES:",
            "• Los nombres de laboratorio deben ser únicos por institución",
            "• Los IDs deben existir en la base de datos",
            "• La capacidad debe ser un número mayor a 0",
            "• La superficie debe ser un número positivo",
            "• El email debe tener formato válido",
            "",
            "🚨 ERRORES COMUNES:",
            "• Nombres de laboratorio duplicados en la misma institución",
            "• IDs que no existen en la base de datos",
            "• Formato incorrecto en campos numéricos",
            "• Emails con formato inválido",
            "",
            "💡 CONSEJOS:",
            "• Revise la hoja 'Referencias' para obtener IDs válidos",
            "• Use la plantilla de ejemplo como guía",
            "• Mantenga el formato de las columnas",
            "• Elimine las filas de ejemplo antes de importar"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row, column=1, value=instruction)
            if instruction.startswith(("📋", "🔴", "🔵", "⚠️", "🚨", "💡")):
                cell.font = Font(bold=True, size=12)
            elif instruction.startswith("•"):
                cell.font = Font(size=10)
        
        # Hoja 3: Referencias
        ws_references = wb.create_sheet("Referencias")
        
        # Instituciones
        instituciones = Institucion.query.all()
        ws_references.cell(row=1, column=1, value="INSTITUCIONES").font = Font(bold=True, size=12)
        ws_references.cell(row=2, column=1, value="ID").font = Font(bold=True)
        ws_references.cell(row=2, column=2, value="Nombre").font = Font(bold=True)
        
        for row, inst in enumerate(instituciones, 3):
            ws_references.cell(row=row, column=1, value=inst.id_institucion)
            ws_references.cell(row=row, column=2, value=inst.nombre_institucion)
        
        # Áreas
        areas = Area.query.all()
        start_row = len(instituciones) + 5
        ws_references.cell(row=start_row, column=1, value="ÁREAS").font = Font(bold=True, size=12)
        ws_references.cell(row=start_row+1, column=1, value="ID").font = Font(bold=True)
        ws_references.cell(row=start_row+1, column=2, value="Nombre").font = Font(bold=True)
        
        for row, area in enumerate(areas, start_row+2):
            ws_references.cell(row=row, column=1, value=area.id_area)
            ws_references.cell(row=row, column=2, value=area.nombre_area)
        
        # Tipos de laboratorio
        tipos = Tipo_Laboratorio.query.all()
        start_row = len(instituciones) + len(areas) + 8
        ws_references.cell(row=start_row, column=1, value="TIPOS DE LABORATORIO").font = Font(bold=True, size=12)
        ws_references.cell(row=start_row+1, column=1, value="ID").font = Font(bold=True)
        ws_references.cell(row=start_row+1, column=2, value="Nombre").font = Font(bold=True)
        
        for row, tipo in enumerate(tipos, start_row+2):
            ws_references.cell(row=row, column=1, value=tipo.id_tipo_laboratorio)
            ws_references.cell(row=row, column=2, value=tipo.nombre_tipo_laboratorio)
        
        # Usuarios (encargados)
        usuarios = User.query.filter(User.activo == True).all()
        start_row = len(instituciones) + len(areas) + len(tipos) + 11
        ws_references.cell(row=start_row, column=1, value="USUARIOS (ENCARGADOS)").font = Font(bold=True, size=12)
        ws_references.cell(row=start_row+1, column=1, value="ID").font = Font(bold=True)
        ws_references.cell(row=start_row+1, column=2, value="Nombre").font = Font(bold=True)
        
        for row, usuario in enumerate(usuarios, start_row+2):
            ws_references.cell(row=row, column=1, value=usuario.id_usuario)
            ws_references.cell(row=row, column=2, value=usuario.nombre_completo)
        
        # Ajustar ancho de columnas
        for ws in [ws_data, ws_instructions, ws_references]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Crear respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=plantilla_laboratorios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        flash(f'Error al generar la plantilla: {str(e)}', 'error')
        return redirect(url_for('admin_laboratorios.admin_laboratorios'))

@admin_laboratorios_bp.route('/admin/laboratorios/importar', methods=['POST'])
@login_required
@permission_required('crear_laboratorio')
def importar_laboratorios():
    """Importar laboratorios desde archivo Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'El archivo debe ser un Excel (.xlsx o .xls)'}), 400
        
        # Leer archivo Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Obtener encabezados
        headers = [cell.value for cell in ws[1]]
        
        # Validar encabezados obligatorios
        campos_obligatorios = ['nombre_laboratorio', 'id_institucion', 'id_tipo_laboratorio']
        headers_lower = [h.lower() if h else '' for h in headers]
        
        for campo in campos_obligatorios:
            if campo not in headers_lower:
                return jsonify({
                    'success': False,
                    'message': f'Falta el campo obligatorio: {campo}'
                }), 400
        
        # Procesar filas
        laboratorios_creados = []
        errores = []
        advertencias = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # Saltar filas vacías
                continue
            
            try:
                # Crear diccionario con los datos de la fila
                data = {}
                for i, header in enumerate(headers):
                    if i < len(row) and header:
                        data[header.lower()] = row[i]
                
                # Validar campos obligatorios
                if not data.get('nombre_laboratorio'):
                    errores.append(f'Fila {row_num}: Nombre del laboratorio es obligatorio')
                    continue
                
                if not data.get('id_institucion'):
                    errores.append(f'Fila {row_num}: ID de institución es obligatorio')
                    continue
                
                if not data.get('id_tipo_laboratorio'):
                    errores.append(f'Fila {row_num}: ID de tipo de laboratorio es obligatorio')
                    continue
                
                # Verificar que no existe un laboratorio con el mismo nombre en la misma institución
                laboratorio_existente = Laboratorio.query.filter_by(
                    nombre_laboratorio=str(data['nombre_laboratorio']).strip(),
                    id_institucion=int(data['id_institucion'])
                ).first()
                
                if laboratorio_existente:
                    errores.append(f'Fila {row_num}: Ya existe un laboratorio con el nombre "{data["nombre_laboratorio"]}" en esta institución')
                    continue
                
                # Validar IDs de referencia
                if not Institucion.query.get(int(data['id_institucion'])):
                    errores.append(f'Fila {row_num}: ID de institución {data["id_institucion"]} no existe')
                    continue
                
                if not Tipo_Laboratorio.query.get(int(data['id_tipo_laboratorio'])):
                    errores.append(f'Fila {row_num}: ID de tipo de laboratorio {data["id_tipo_laboratorio"]} no existe')
                    continue
                
                # Validar área si se proporciona
                if data.get('id_area') and data['id_area']:
                    if not Area.query.get(int(data['id_area'])):
                        errores.append(f'Fila {row_num}: ID de área {data["id_area"]} no existe')
                        continue
                
                # Validar encargado si se proporciona
                if data.get('id_encargado') and data['id_encargado']:
                    if not User.query.get(int(data['id_encargado'])):
                        errores.append(f'Fila {row_num}: ID de encargado {data["id_encargado"]} no existe')
                        continue
                
                # Crear laboratorio
                nuevo_laboratorio = Laboratorio(
                    nombre_laboratorio=str(data['nombre_laboratorio']).strip(),
                    descripcion=str(data.get('descripcion', '')).strip(),
                    disponibilidad=str(data.get('disponibilidad', 'Activo')).strip(),
                    capacidad=int(data.get('capacidad', 20)),
                    horario=str(data.get('horario', '')).strip(),
                    ubicacion=str(data.get('ubicacion', '')).strip(),
                    telefono=str(data.get('telefono', '')).strip(),
                    email_contacto=str(data.get('email_contacto', '')).strip(),
                    superficie_m2=float(data.get('superficie_m2', 0)) if data.get('superficie_m2') else None,
                    equipamiento=str(data.get('equipamiento', '')).strip(),
                    normas_uso=str(data.get('normas_uso', '')).strip(),
                    requiere_capacitacion=str(data.get('requiere_capacitacion', 'No')).strip().lower() in ['sí', 'si', 'yes', 'true', '1'],
                    id_institucion=int(data['id_institucion']),
                    id_area=int(data['id_area']) if data.get('id_area') and data['id_area'] else None,
                    id_tipo_laboratorio=int(data['id_tipo_laboratorio']),
                    id_encargado=int(data['id_encargado']) if data.get('id_encargado') and data['id_encargado'] else None,
                    fecha_creacion=datetime.utcnow(),
                    fecha_modificacion=datetime.utcnow()
                )
                
                db.session.add(nuevo_laboratorio)
                laboratorios_creados.append(nuevo_laboratorio.nombre_laboratorio)
                
            except ValueError as e:
                errores.append(f'Fila {row_num}: Error en formato de datos - {str(e)}')
            except Exception as e:
                errores.append(f'Fila {row_num}: Error inesperado - {str(e)}')
        
        # Confirmar transacción si no hay errores críticos
        if laboratorios_creados:
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Importación completada. {len(laboratorios_creados)} laboratorios creados.',
            'laboratorios_creados': laboratorios_creados,
            'errores': errores,
            'advertencias': advertencias
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al importar archivo: {str(e)}'
        }), 500

@admin_laboratorios_bp.route('/admin/laboratorios/buscar')
@login_required
@permission_required('gestionar_laboratorios')
def buscar_laboratorios():
    """Buscar laboratorios para autocompletado"""
    try:
        query = request.args.get('q', '').strip()
        
        if not query:
            return jsonify({'success': True, 'laboratorios': []})
        
        laboratorios = Laboratorio.query.filter(
            or_(
                Laboratorio.nombre_laboratorio.ilike(f'%{query}%'),
                Laboratorio.descripcion.ilike(f'%{query}%')
            )
        ).limit(10).all()
        
        return jsonify({
            'success': True,
            'laboratorios': [{
                'id': lab.id_laboratorio,
                'nombre': lab.nombre_laboratorio,
                'descripcion': lab.descripcion,
                'disponibilidad': lab.disponibilidad
            } for lab in laboratorios]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error en la búsqueda: {str(e)}'
        }), 500 