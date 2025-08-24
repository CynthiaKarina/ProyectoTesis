from flask import Blueprint, render_template, session, redirect, url_for, flash, request, jsonify, make_response
from flask_login import login_required
from app.models.user import User
from app.models.area import Area
from app.models.institucion import Institucion
from app.models.tipo_institucion import Tipo_Institucion
from app.models.roles import Roles
from app import db
from datetime import datetime
from app.utils.permissions import permission_required, any_permission_required
import io
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

admin_users_bp = Blueprint('admin_users', __name__)
@admin_users_bp.route('/admin/usuarios/inactividad/ejecutar', methods=['POST'])
@login_required
@permission_required('gestionar_usuarios')
def ejecutar_inactividad():
    try:
        from app.utils.inactivity import process_inactive_accounts_from_env
        result = process_inactive_accounts_from_env()
        flash(f"Inactividad procesada. Advertidos: {result.get('warned',0)}, Eliminados: {result.get('deleted',0)}", 'success')
    except Exception as e:
        flash(f'Error ejecutando proceso de inactividad: {e}', 'error')
    return redirect(url_for('admin_users.admin_usuarios'))

@admin_users_bp.route('/admin/usuarios')
@login_required
@permission_required('gestionar_usuarios')
def admin_usuarios():
    usuarios = User.query.all()
    areas = Area.query.all()
    instituciones = Institucion.query.all()
    roles = Roles.query.all()
    return render_template('admin_usuarios.html', usuarios=usuarios, areas=areas, instituciones=instituciones, roles=roles)

@admin_users_bp.route('/admin/usuarios/agregar', methods=['POST'])
@login_required
@permission_required('crear_usuario')
def agregar_usuario():
    try:
        data = request.form
        nuevo_usuario = User(
            nombre=data.get('nombre'),
            apellido_paterno=data.get('apellido_paterno'),
            apellido_materno=data.get('apellido_materno'),
            email=data.get('email'),
            username=data.get('username'),
            telefono=data.get('telefono'),
            id_institucion=data.get('id_institucion'),
            id_rol=data.get('id_rol'),
            activo=bool(int(data.get('activo', 1))),
            id_area=data.get('id_area')
        )
        db.session.add(nuevo_usuario)
        db.session.commit()
        return jsonify({'mensaje': 'Usuario agregado exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@admin_users_bp.route('/admin/usuarios/editar/<int:id>', methods=['PUT'])
@login_required
@permission_required('editar_usuario')
def editar_usuario(id):
    try:
        usuario = User.query.get_or_404(id)
        
        # Verificar si el usuario es Super Administrador (rol_id = 16)
        if usuario.id_rol == 16:
            return jsonify({'error': 'No se puede editar un usuario Super Administrador'}), 403
        
        data = request.get_json()
        usuario.nombre = data.get('nombre', usuario.nombre)
        usuario.apellido_paterno = data.get('apellido_paterno', usuario.apellido_paterno)
        usuario.apellido_materno = data.get('apellido_materno', usuario.apellido_materno)
        usuario.email = data.get('email', usuario.email)
        usuario.username = data.get('username', usuario.username)
        usuario.telefono = data.get('telefono', usuario.telefono)
        usuario.id_institucion = data.get('id_institucion', usuario.id_institucion)
        usuario.id_rol = data.get('id_rol', usuario.id_rol)
        usuario.activo = bool(int(data.get('activo', usuario.activo)))
        usuario.id_area = data.get('id_area', usuario.id_area)
        db.session.commit()
        return jsonify({'mensaje': 'Usuario actualizado exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@admin_users_bp.route('/admin/usuarios/eliminar/<int:id>', methods=['DELETE'])
@login_required
@permission_required('eliminar_usuario')
def eliminar_usuario(id):
    try:
        usuario = User.query.get_or_404(id)
        db.session.delete(usuario)
        db.session.commit()
        return jsonify({'mensaje': 'Usuario eliminado exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@admin_users_bp.route('/admin/usuarios/verificar/<email>/<username>/<id_usuario>')
@login_required
@any_permission_required(['gestionar_usuarios', 'crear_usuario', 'editar_usuario'])
def verificar_usuario(email, username, id_usuario):
    try:
        # Si es una edición, excluir el usuario actual de la búsqueda
        query = User.query
        if id_usuario:
            query = query.filter(User.id_usuario != id_usuario)
        
        # Verificar email
        email_existe = query.filter(User.email == email).first() is not None
        # Verificar username
        username_existe = query.filter(User.username == username).first() is not None
        
        return jsonify({
            'existe': email_existe or username_existe,
            'email_existe': email_existe,
            'username_existe': username_existe
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_users_bp.route('/admin/usuarios/detalles/<int:id>')
@login_required
@permission_required('gestionar_usuarios')
def obtener_detalles_usuario(id):
    try:
        usuario = User.query.get_or_404(id)
        # Manejar fechas que pueden ser string o None
        def formatear_fecha(fecha, formato):
            if fecha is None:
                return 'No disponible'
            if hasattr(fecha, 'strftime'):
                return fecha.strftime(formato)
            return str(fecha)
        return jsonify({
            'id': usuario.id_usuario,
            'estado': 'Activo' if usuario.activo else 'Inactivo',
            'ultima_conexion': formatear_fecha(usuario.ultimo_acceso, '%d/%m/%Y %H:%M'),
            'fecha_registro': formatear_fecha(usuario.fecha_registro, '%d/%m/%Y')
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@admin_users_bp.route('/admin/usuarios/actualizar_rol', methods=['POST'])
@login_required
@permission_required('editar_usuario')
def actualizar_rol_usuario():
    user_id = request.form.get('user_id')
    id_rol = request.form.get('id_rol')
    usuario = User.query.get(user_id)
    if not usuario:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    rol_anterior = usuario.id_rol
    usuario.id_rol = id_rol
    db.session.commit()
    # Auditoría básica en DB y logger
    try:
        from flask_login import current_user
        from flask import current_app
        from app.models.audit_log import AuditLog
        db.session.add(AuditLog(
            action='cambio_rol_usuario',
            actor_id=getattr(current_user, 'id_usuario', None),
            target_user_id=usuario.id_usuario,
            from_value=str(rol_anterior),
            to_value=str(id_rol),
            extra=None
        ))
        db.session.commit()
        current_app.logger.info(
            f"[AUDITORIA ROL] actor_id={getattr(current_user, 'id_usuario', None)} usuario_id={usuario.id_usuario} rol_anterior={rol_anterior} rol_nuevo={id_rol} fecha={datetime.now().isoformat()}"
        )
    except Exception:
        pass
    return jsonify({'success': True, 'mensaje': 'Rol actualizado'})

@admin_users_bp.route('/admin/usuarios/descargar-instituciones')
@login_required
@permission_required('gestionar_usuarios')
def descargar_instituciones():
    """
    Descarga la lista actualizada de instituciones en formato CSV
    """
    try:
        # Obtener todas las instituciones (primero intentar con JOIN)
        instituciones_con_tipo = db.session.query(Institucion, Tipo_Institucion).join(
            Tipo_Institucion, Institucion.id_tipo_institucion == Tipo_Institucion.id_tipo_institucion
        ).filter(Institucion.activo == True).all()
        
        # Si no hay resultados con JOIN, obtener todas las instituciones activas
        if not instituciones_con_tipo:
            instituciones_sin_tipo = Institucion.query.filter_by(activo=True).all()
            # Convertir a formato compatible
            instituciones_con_tipo = []
            for inst in instituciones_sin_tipo:
                tipo_inst = None
                if inst.id_tipo_institucion:
                    tipo_inst = Tipo_Institucion.query.get(inst.id_tipo_institucion)
                
                # Crear un objeto tipo dummy si no existe
                if not tipo_inst:
                    class TipoDummy:
                        def __init__(self):
                            self.nombre_tipo_institucion = "Sin tipo definido"
                    tipo_inst = TipoDummy()
                
                instituciones_con_tipo.append((inst, tipo_inst))
        
        # Crear el archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow([
            'ID',
            'Nombre de la Institución',
            'Tipo de Institución',
            'Dirección Completa',
            'Teléfono',
            'Email',
            'Estado',
            'Fecha de Registro'
        ])
        
        # Si no hay instituciones, agregar una fila indicándolo
        if not instituciones_con_tipo:
            writer.writerow([
                'N/A',
                'No hay instituciones registradas',
                'N/A',
                'N/A',
                'N/A',
                'N/A',
                'N/A',
                'N/A'
            ])
        else:
            # Escribir datos de instituciones
            for institucion, tipo_institucion in instituciones_con_tipo:
                # Construir dirección completa
                direccion_parts = []
                if institucion.calle:
                    direccion_parts.append(institucion.calle)
                if institucion.colonia:
                    direccion_parts.append(f"Col. {institucion.colonia}")
                if institucion.municipio:
                    direccion_parts.append(institucion.municipio)
                if institucion.estado:
                    direccion_parts.append(institucion.estado)
                if institucion.codigo_postal:
                    direccion_parts.append(f"C.P. {institucion.codigo_postal}")
                
                direccion_completa = ", ".join(direccion_parts) if direccion_parts else "Sin dirección"
                
                # Formatear fecha de registro de manera segura
                fecha_registro_str = 'Sin fecha'
                if institucion.fecha_registro:
                    try:
                        if hasattr(institucion.fecha_registro, 'strftime'):
                            fecha_registro_str = institucion.fecha_registro.strftime('%d/%m/%Y %H:%M')
                        else:
                            # Si es string, intentar convertir a datetime
                            from datetime import datetime
                            if isinstance(institucion.fecha_registro, str):
                                fecha_obj = datetime.fromisoformat(institucion.fecha_registro.replace('Z', '+00:00'))
                                fecha_registro_str = fecha_obj.strftime('%d/%m/%Y %H:%M')
                            else:
                                fecha_registro_str = str(institucion.fecha_registro)
                    except Exception as e:
                        fecha_registro_str = f'Fecha inválida: {str(institucion.fecha_registro)}'
                
                writer.writerow([
                    institucion.id_institucion,
                    institucion.nombre_institucion,
                    tipo_institucion.nombre_tipo_institucion,
                    direccion_completa,
                    institucion.telefono or 'Sin teléfono',
                    institucion.email or 'Sin email',
                    'Activa' if institucion.activo else 'Inactiva',
                    fecha_registro_str
                ])
        
        # Preparar la respuesta
        output.seek(0)
        csv_content = output.getvalue()
        
        # Crear respuesta con el archivo CSV
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=instituciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        # Log del error para debugging
        print(f"Error en descargar_instituciones: {str(e)}")
        
        # Crear un CSV de error
        error_csv = f"Error,{str(e)}\n"
        response = make_response(error_csv)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=error_instituciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response


@admin_users_bp.route('/admin/usuarios/exportar')
@login_required
@permission_required('gestionar_usuarios')
def exportar_usuarios():
    """Exportar usuarios en formato CSV"""
    try:
        # Crear un buffer de memoria para el CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow([
            'ID',
            'Nombre',
            'Apellido Paterno',
            'Apellido Materno',
            'Email',
            'Username',
            'Teléfono',
            'Institución',
            'Área',
            'Rol',
            'Estado',
            'Fecha de Registro'
        ])
        
        # Obtener usuarios con sus relaciones
        usuarios = db.session.query(
            User, Institucion, Area, Roles
        ).outerjoin(
            Institucion, User.id_institucion == Institucion.id_institucion
        ).outerjoin(
            Area, User.id_area == Area.id_area
        ).outerjoin(
            Roles, User.id_rol == Roles.id_rol
        ).all()
        
        # Verificar si hay datos
        if not usuarios:
            writer.writerow([
                'N/A',
                'No hay usuarios registrados',
                'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A'
            ])
        else:
            # Escribir datos de usuarios
            for usuario, institucion, area, rol in usuarios:
                # Formatear fecha de registro de manera segura
                fecha_registro_str = 'Sin fecha'
                if usuario.fecha_registro:
                    try:
                        if hasattr(usuario.fecha_registro, 'strftime'):
                            fecha_registro_str = usuario.fecha_registro.strftime('%d/%m/%Y %H:%M')
                        else:
                            fecha_registro_str = str(usuario.fecha_registro)
                    except Exception as e:
                        fecha_registro_str = f'Fecha inválida: {str(usuario.fecha_registro)}'
                
                writer.writerow([
                    usuario.id_usuario,
                    usuario.nombre or 'Sin nombre',
                    usuario.apellido_paterno or 'Sin apellido paterno',
                    usuario.apellido_materno or 'Sin apellido materno',
                    usuario.email or 'Sin email',
                    usuario.username or 'Sin username',
                    usuario.telefono or 'Sin teléfono',
                    institucion.nombre_institucion if institucion else 'Sin institución',
                    area.nombre_area if area else 'Sin área',
                    rol.nombre_rol if rol else 'Sin rol',
                    'Activo' if usuario.activo else 'Inactivo',
                    fecha_registro_str
                ])
        
        # Preparar la respuesta
        output.seek(0)
        csv_content = output.getvalue()
        
        # Crear respuesta con el archivo CSV
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        # Log del error para debugging
        print(f"Error en exportar_usuarios: {str(e)}")
        
        # Crear un CSV de error
        error_csv = f"Error,{str(e)}\n"
        response = make_response(error_csv)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=error_usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response


@admin_users_bp.route('/admin/usuarios/plantilla')
@login_required
@permission_required('crear_usuario')
def descargar_plantilla():
    """Descargar plantilla de Excel para importar usuarios"""
    try:
        # Crear un nuevo workbook de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Usuarios"
        
        # Definir encabezados
        encabezados = [
            'nombre',
            'apellido_paterno', 
            'apellido_materno',
            'email',
            'username',
            'telefono',
            'id_institucion',
            'id_area',
            'id_rol',
            'activo'
        ]
        
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")  # Rojo
        optional_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")  # Azul
        example_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # Gris claro
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Campos obligatorios (primeros 5)
        campos_obligatorios = ['nombre', 'apellido_paterno', 'email', 'username', 'id_institucion']
        
        # Escribir encabezados con formato
        for col, header in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            
                         # Colorear según si es obligatorio o opcional
            if header in campos_obligatorios:
                cell.fill = required_fill
            else:
                cell.fill = optional_fill
        
        # Agregar fila explicativa
        explicaciones = [
            'OBLIGATORIO',      # nombre
            'OBLIGATORIO',      # apellido_paterno
            'OPCIONAL',         # apellido_materno
            'OBLIGATORIO',      # email
            'OBLIGATORIO',      # username
            'OPCIONAL',         # telefono
            'OBLIGATORIO',      # id_institucion
            'OPCIONAL',         # id_area
            'OPCIONAL',         # id_rol
            'OPCIONAL'          # activo
        ]
        
        for col, explicacion in enumerate(explicaciones, 1):
            cell = ws.cell(row=2, column=col, value=explicacion)
            cell.font = Font(bold=True, size=9)
            cell.alignment = center_alignment
            cell.border = border
            if explicacion == 'OBLIGATORIO':
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Agregar ejemplos más detallados (ahora en fila 3 en adelante)
        ejemplos = [
            ['Juan', 'Pérez', 'García', 'juan.perez@universidad.edu', 'jperez', '5551234567', '1', '1', '2', '1'],
            ['María', 'López', 'Martínez', 'maria.lopez@instituto.edu', 'mlopez', '5555678901', '2', '2', '3', '1'],
            ['Carlos', 'Rodríguez', 'Hernández', 'carlos.rodriguez@colegio.edu', 'crodriguez', '5559012345', '1', '2', '2', '1'],
            ['Ana', 'González', 'Morales', 'ana.gonzalez@universidad.edu', 'agonzalez', '5556789012', '3', '1', '8', '1'],
            ['Luis', 'Martínez', '', 'luis.martinez@instituto.edu', 'lmartinez', '', '1', '3', '9', '0']
        ]
        
        for row_num, ejemplo in enumerate(ejemplos, 3):  # Empezar en fila 3
            for col, valor in enumerate(ejemplo, 1):
                cell = ws.cell(row=row_num, column=col, value=valor)
                cell.fill = example_fill
                cell.border = border
                cell.alignment = center_alignment
        
        # Ajustar ancho de columnas
        column_widths = [15, 18, 18, 30, 15, 15, 15, 12, 12, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Crear una segunda hoja con instrucciones
        ws_instrucciones = wb.create_sheet(title="Instrucciones")
        
        instrucciones = [
            ["📋 GUÍA COMPLETA PARA IMPORTAR USUARIOS"],
            [""],
            ["🔴 CAMPOS OBLIGATORIOS (REQUERIDOS):"],
            ["✅ nombre: Nombre del usuario (ej: Juan, María, Carlos)"],
            ["✅ apellido_paterno: Apellido paterno (ej: Pérez, López, García)"],
            ["✅ email: Correo electrónico único (ej: juan.perez@universidad.edu)"],
            ["✅ username: Nombre de usuario único (ej: jperez, mlopez, cgarcia)"],
            ["✅ id_institucion: ID de la institución (OBLIGATORIO - ver hoja 'IDs de Referencia')"],
            [""],
            ["🔵 CAMPOS OPCIONALES:"],
            ["• apellido_materno: Apellido materno del usuario"],
            ["• telefono: Número de teléfono (ej: 5551234567)"],
            ["• id_area: ID del área académica (consultar hoja 'IDs de Referencia')"],
            ["• id_rol: ID del rol del usuario (consultar hoja 'IDs de Referencia')"],
            ["• activo: 1 para activo, 0 para inactivo (por defecto: 1)"],
            [""],
            ["⚠️ VALIDACIONES IMPORTANTES:"],
            ["• El email debe ser único en el sistema (no se permiten duplicados)"],
            ["• Si el username ya existe, se generará uno automáticamente"],
            ["• El ID de institución es OBLIGATORIO y debe existir en el sistema"],
            ["• Los IDs de área y rol deben existir en el sistema (si se proporcionan)"],
            ["• El email debe tener formato válido (debe contener @ y .)"],
            ["• Los campos obligatorios no pueden estar vacíos"],
            [""],
            ["📌 EJEMPLOS DE DATOS VÁLIDOS:"],
            ["• Email: usuario@dominio.com, estudiante@universidad.edu"],
            ["• Username: jperez, maria.lopez, carlos123"],
            ["• Teléfono: 5551234567, 555-123-4567"],
            ["• Estado activo: 1 (activo), 0 (inactivo)"],
            [""],
            ["🔗 OBTENER IDs DE REFERENCIA:"],
            ["• Consulta la hoja 'IDs de Referencia' en este archivo"],
            ["• Descarga la lista actualizada desde la página de administración"],
            ["• Usa los botones 'Ver IDs Disponibles' en el sistema"],
            [""],
            ["❌ ERRORES COMUNES A EVITAR:"],
            ["• Dejar campos obligatorios vacíos (nombre, apellido_paterno, email, username, id_institucion)"],
            ["• Usar emails duplicados"],
            ["• No proporcionar ID de institución (campo obligatorio)"],
            ["• IDs de institución/área/rol que no existen en el sistema"],
            ["• Formato de email inválido"],
            ["• Caracteres especiales en username"]
        ]
        
        for row_num, instruccion in enumerate(instrucciones, 1):
            ws_instrucciones.cell(row=row_num, column=1, value=instruccion[0])
        
        # Crear una tercera hoja con IDs de referencia
        ws_referencia = wb.create_sheet(title="IDs de Referencia")
        
        # Obtener datos de referencia
        instituciones = Institucion.query.filter_by(activo=True).all()
        areas = Area.query.all()
        roles = Roles.query.all()
        
        # Escribir instituciones
        ws_referencia.cell(row=1, column=1, value="INSTITUCIONES")
        ws_referencia.cell(row=2, column=1, value="ID")
        ws_referencia.cell(row=2, column=2, value="Nombre")
        
        for i, inst in enumerate(instituciones, 3):
            ws_referencia.cell(row=i, column=1, value=inst.id_institucion)
            ws_referencia.cell(row=i, column=2, value=inst.nombre_institucion)
        
        # Escribir áreas
        start_row = len(instituciones) + 5
        ws_referencia.cell(row=start_row, column=1, value="ÁREAS")
        ws_referencia.cell(row=start_row + 1, column=1, value="ID")
        ws_referencia.cell(row=start_row + 1, column=2, value="Nombre")
        
        for i, area in enumerate(areas, start_row + 2):
            ws_referencia.cell(row=i, column=1, value=area.id_area)
            ws_referencia.cell(row=i, column=2, value=area.nombre_area)
        
        # Escribir roles
        start_row = start_row + len(areas) + 3
        ws_referencia.cell(row=start_row, column=1, value="ROLES")
        ws_referencia.cell(row=start_row + 1, column=1, value="ID")
        ws_referencia.cell(row=start_row + 1, column=2, value="Nombre")
        
        for i, rol in enumerate(roles, start_row + 2):
            ws_referencia.cell(row=i, column=1, value=rol.id_rol)
            ws_referencia.cell(row=i, column=2, value=rol.nombre_rol)
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=plantilla_usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        # Log del error para debugging
        print(f"Error en descargar_plantilla: {str(e)}")
        
        # Crear un archivo de error simple
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Error")
        ws.cell(row=1, column=2, value=str(e))
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=error_plantilla_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response


@admin_users_bp.route('/admin/usuarios/importar', methods=['POST'])
@login_required
@permission_required('crear_usuario')
def importar_usuarios():
    """Importar usuarios desde archivo CSV/Excel con validaciones completas"""
    try:
        # Verificar que se haya enviado un archivo
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        archivo = request.files['archivo']
        
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        # Verificar extensión del archivo
        extension = archivo.filename.lower().split('.')[-1]
        if extension not in ['csv', 'xlsx', 'xls']:
            return jsonify({'error': 'Formato de archivo no válido. Use CSV o Excel (.xlsx, .xls)'}), 400
        
        # Procesar archivo según su tipo
        datos = []
        
        if extension == 'csv':
            # Procesar CSV
            contenido = archivo.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(contenido))
            datos = list(reader)
        else:
            # Procesar Excel
            workbook = openpyxl.load_workbook(archivo)
            sheet = workbook.active
            
            # Obtener encabezados (primera fila)
            encabezados = []
            for cell in sheet[1]:
                if cell.value:
                    encabezados.append(str(cell.value).strip())
            
            # Procesar datos
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # Si la fila tiene al menos un valor
                    fila_dict = {}
                    for i, valor in enumerate(row):
                        if i < len(encabezados):
                            fila_dict[encabezados[i]] = str(valor).strip() if valor is not None else ''
                    datos.append(fila_dict)
        
        # Validar y procesar datos
        usuarios_creados = 0
        usuarios_actualizados = 0
        errores = []
        warnings = []
        
        # Campos obligatorios
        campos_obligatorios = ['nombre', 'apellido_paterno', 'email', 'username', 'id_institucion']
        
        for fila_num, fila in enumerate(datos, start=2):  # Empezar en 2 porque la fila 1 son encabezados
            try:
                # Validar campos obligatorios
                campos_faltantes = []
                for campo in campos_obligatorios:
                    if not fila.get(campo) or fila.get(campo).strip() == '':
                        campos_faltantes.append(campo)
                
                if campos_faltantes:
                    errores.append(f'Fila {fila_num}: Faltan campos obligatorios: {", ".join(campos_faltantes)}')
                    continue
                
                # Limpiar y validar datos
                email = fila.get('email').strip().lower()
                username = fila.get('username').strip()
                nombre = fila.get('nombre').strip()
                apellido_paterno = fila.get('apellido_paterno').strip()
                
                # Validar formato de email básico
                if '@' not in email or '.' not in email:
                    errores.append(f'Fila {fila_num}: Email inválido: {email}')
                    continue
                
                # Verificar si el usuario ya existe (por email)
                usuario_existente = User.query.filter_by(email=email).first()
                
                if usuario_existente:
                    warnings.append(f'Fila {fila_num}: Usuario con email {email} ya existe - OMITIDO')
                    continue
                
                # Verificar si el username ya existe
                username_existente = User.query.filter_by(username=username).first()
                if username_existente:
                    # Generar username único
                    contador = 1
                    username_original = username
                    while User.query.filter_by(username=username).first():
                        username = f"{username_original}_{contador}"
                        contador += 1
                    warnings.append(f'Fila {fila_num}: Username modificado de {username_original} a {username}')
                
                # Validar IDs de relaciones
                id_institucion = None
                id_area = None
                id_rol = None
                
                # Validar institución (OBLIGATORIO)
                if not fila.get('id_institucion') or not fila.get('id_institucion').strip():
                    errores.append(f'Fila {fila_num}: ID de institución es obligatorio')
                    continue
                
                try:
                    id_institucion = int(fila.get('id_institucion'))
                    if not Institucion.query.get(id_institucion):
                        errores.append(f'Fila {fila_num}: Institución con ID {id_institucion} no existe')
                        continue
                except ValueError:
                    errores.append(f'Fila {fila_num}: ID de institución inválido: {fila.get("id_institucion")}')
                    continue
                
                # Validar área
                if fila.get('id_area') and fila.get('id_area').strip():
                    try:
                        id_area = int(fila.get('id_area'))
                        if not Area.query.get(id_area):
                            errores.append(f'Fila {fila_num}: Área con ID {id_area} no existe')
                            continue
                    except ValueError:
                        errores.append(f'Fila {fila_num}: ID de área inválido: {fila.get("id_area")}')
                        continue
                
                # Validar rol
                if fila.get('id_rol') and fila.get('id_rol').strip():
                    try:
                        id_rol = int(fila.get('id_rol'))
                        if not Roles.query.get(id_rol):
                            errores.append(f'Fila {fila_num}: Rol con ID {id_rol} no existe')
                            continue
                    except ValueError:
                        errores.append(f'Fila {fila_num}: ID de rol inválido: {fila.get("id_rol")}')
                        continue
                
                # Validar estado activo
                activo = True  # Por defecto activo
                if fila.get('activo') and fila.get('activo').strip():
                    activo_str = fila.get('activo').strip().lower()
                    if activo_str in ['0', 'false', 'no', 'inactivo']:
                        activo = False
                    elif activo_str in ['1', 'true', 'si', 'sí', 'activo']:
                        activo = True
                    else:
                        warnings.append(f'Fila {fila_num}: Valor de activo inválido "{fila.get("activo")}", usando "activo" por defecto')
                
                # Crear nuevo usuario
                nuevo_usuario = User(
                    nombre=nombre,
                    apellido_paterno=apellido_paterno,
                    apellido_materno=fila.get('apellido_materno', '').strip() if fila.get('apellido_materno') else None,
                    email=email,
                    username=username,
                    telefono=fila.get('telefono', '').strip() if fila.get('telefono') else None,
                    id_institucion=id_institucion,
                    id_area=id_area,
                    id_rol=id_rol,
                    activo=activo,
                    fecha_registro=datetime.now()
                )
                
                db.session.add(nuevo_usuario)
                usuarios_creados += 1
                
            except Exception as e:
                errores.append(f'Fila {fila_num}: Error procesando datos - {str(e)}')
        
        # Confirmar cambios si hay usuarios para crear
        if usuarios_creados > 0:
            db.session.commit()
        
        # Preparar respuesta detallada
        respuesta = {
            'usuarios_creados': usuarios_creados,
            'usuarios_actualizados': usuarios_actualizados,
            'total_procesados': len(datos),
            'errores': errores,
            'warnings': warnings,
            'mensaje': f'Importación completada: {usuarios_creados} usuarios creados'
        }
        
        if errores:
            respuesta['mensaje'] += f', {len(errores)} errores encontrados'
        
        if warnings:
            respuesta['mensaje'] += f', {len(warnings)} advertencias'
        
        return jsonify(respuesta), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error en importación: {str(e)}'}), 500


# Vista simple de auditoría de cambios (roles y futuras acciones)
@admin_users_bp.route('/admin/auditoria')
@login_required
@permission_required('gestionar_usuarios')
def ver_auditoria():
    try:
        from app.models.audit_log import AuditLog
        # Filtros
        action = request.args.get('action', '').strip()
        actor_id = request.args.get('actor_id', '').strip()
        target_user_id = request.args.get('target_user_id', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        page = max(int(request.args.get('page', 1) or 1), 1)
        per_page = max(min(int(request.args.get('per_page', 20) or 20), 100), 5)

        q = db.session.query(AuditLog)
        if action:
            q = q.filter(AuditLog.action == action)
        if actor_id:
            try:
                q = q.filter(AuditLog.actor_id == int(actor_id))
            except Exception:
                pass
        if target_user_id:
            try:
                q = q.filter(AuditLog.target_user_id == int(target_user_id))
            except Exception:
                pass
        # Fechas
        from datetime import datetime, timedelta
        def parse_date(d):
            try:
                return datetime.strptime(d, '%Y-%m-%d')
            except Exception:
                return None
        if date_from:
            df = parse_date(date_from)
            if df:
                q = q.filter(AuditLog.created_at >= df)
        if date_to:
            dt = parse_date(date_to)
            if dt:
                q = q.filter(AuditLog.created_at < (dt + timedelta(days=1)))

        total = q.count()
        q = q.order_by(AuditLog.created_at.desc())
        logs = q.offset((page - 1) * per_page).limit(per_page).all()

        # Mapear usuarios actor/target
        actor_ids = {l.actor_id for l in logs if getattr(l, 'actor_id', None)}
        target_ids = {l.target_user_id for l in logs if getattr(l, 'target_user_id', None)}
        all_ids = sorted(actor_ids.union(target_ids))
        users_map = {}
        if all_ids:
            users = db.session.query(User).filter(User.id_usuario.in_(all_ids)).all()
            for u in users:
                users_map[u.id_usuario] = f"{u.nombre or u.username} ({u.email})"

        # Acciones distintas para el filtro
        actions = [row[0] for row in db.session.query(AuditLog.action).distinct().all()]

        total_pages = (total + per_page - 1) // per_page
        return render_template(
            'admin_auditoria.html',
            logs=logs,
            users_map=users_map,
            actions=actions,
            total=total,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            f_action=action,
            f_actor_id=actor_id,
            f_target_user_id=target_user_id,
            f_date_from=date_from,
            f_date_to=date_to,
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
